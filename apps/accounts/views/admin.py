import json
import logging

from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.cache import cache
from django.conf import settings
from apps.accounts.decorators import role_required
from django.http import HttpResponse, Http404
from django.core.exceptions import PermissionDenied, ValidationError
from django.db.models import Count, Q, Sum, F, FloatField, ExpressionWrapper, Avg
from django.core.paginator import Paginator
from datetime import date, timedelta
from django.utils import timezone

from apps.accounts.models import User, TeacherAbsence, TeacherSubstitution, Batch
from apps.accounts import scoping
from apps.accounts import scoping
from apps.accounts.scoping import scoped_batch as _sub_admin_batch
from apps.accounts.scoping import check_batch_access as _check_batch_access
from apps.accounts.scoping import scoped_batch_ids as _scoped_batch_ids
from apps.accounts.forms import SignupForm, ApprovalForm, BatchForm, ProfileForm
from apps.accounts.utils.email import send_approval_email, send_rejection_email

logger = logging.getLogger(__name__)

from apps.circles.models import Circle, CircleEnrollment, Session
from apps.attendance.models import Attendance
from apps.requests.models import SupportRequest
from apps.announcements.models import Announcement
from apps.notifications.models import Notification
from apps.memorization.models import MemorizationProgress, RecitationGrade
from apps.exams.models import Exam, ExamMark
from apps.certificates.models import Certificate
from apps.references.models import Surah, EvaluationCriterion
from apps.references.utils import (
    TOTAL_THUMNS, count_thumns, format_hizb_thumn, thumn_start_keys,
)

@login_required
@role_required(User.Role.MAIN_ADMIN, User.Role.SUB_ADMIN)
def admin_dashboard(request):

    total_students = User.objects.filter(
        role=User.Role.STUDENT, is_approved=User.ApprovalStatus.APPROVED
    ).count()
    total_teachers = User.objects.filter(
        role=User.Role.TEACHER, is_approved=User.ApprovalStatus.APPROVED
    ).count()
    active_circles = Circle.objects.filter(status=Circle.Status.ACTIVE).count()
    total_supervisors = User.objects.filter(role=User.Role.SUB_ADMIN).count()

    circles_dist = (
        Circle.objects.annotate(count=Count('enrollments', filter=Q(enrollments__status=CircleEnrollment.Status.ACTIVE)))
        .order_by('-count')[:5]
    )
    max_count = max(max((c.count for c in circles_dist), default=0), 1)
    circle_distribution = [
        {'name': c.name, 'count': c.count, 'percentage': round(c.count / max_count * 100)}
        for c in circles_dist
    ]

    total_att = cache.get_or_set("dash:total_att", Attendance.objects.count, 120)
    present_att = cache.get_or_set(
        "dash:present_att",
        lambda: Attendance.objects.filter(
            status__in=[Attendance.Status.PRESENT, Attendance.Status.LATE, Attendance.Status.LEFT_EARLY]
        ).count(),
        120,
    )
    attendance_rate = round(present_att / total_att * 100, 1) if total_att else 0
    total_absent = cache.get_or_set(
        "dash:absent_att",
        lambda: Attendance.objects.filter(
            status__in=[Attendance.Status.ABSENT_UNJUSTIFIED, Attendance.Status.ABSENT, Attendance.Status.ABSENT_JUSTIFIED]
        ).count(),
        120,
    )

    # Memorization stats in the platform tracking unit (thumn/hizb)
    total_hifz_thumns = cache.get_or_set(
        "dash:hifz_thumns",
        lambda: count_thumns(
            MemorizationProgress.objects.filter(type='hifz')
            .values_list('surah_id', 'ayah_from', 'ayah_to')
        ),
        120,
    )
    total_murajaa_thumns = cache.get_or_set(
        "dash:murajaa_thumns",
        lambda: count_thumns(
            MemorizationProgress.objects.filter(type='murajaa')
            .values_list('surah_id', 'ayah_from', 'ayah_to')
        ),
        120,
    )

    day_names_ar = ['السبت', 'الأحد', 'الإثنين', 'الثلاثاء', 'الأربعاء', 'الخميس', 'الجمعة']
    weekly_attendance = []
    for i in range(6, -1, -1):
        d = date.today() - timedelta(days=i)
        day_total = Attendance.objects.filter(session__session_date=d).count()
        day_present = Attendance.objects.filter(
            session__session_date=d, status__in=['present', 'late']
        ).count()
        pct = round(day_present / day_total * 100) if day_total else 0
        weekly_attendance.append({'label': day_names_ar[d.weekday()], 'percentage': pct})

    recent_users = User.objects.order_by('-created_at')[:5]
    pending_users = User.objects.filter(is_approved='pending')[:5]
    pending_count = User.objects.filter(is_approved='pending').count()
    total_certificates = Certificate.objects.filter(status="issued").count()
    approved_count = User.objects.filter(is_approved='approved').count()
    rejected_count = User.objects.filter(is_approved='rejected').count()

    # Recent requests
    recent_requests = SupportRequest.objects.select_related('submitted_by').order_by('-created_at')[:5]
    urgent_requests = SupportRequest.objects.filter(priority__in=['urgent', 'high'], status__in=['submitted', 'under_review']).count()

    today = date.today()
    today_sessions = Session.objects.filter(session_date=today).count()
    today_attendance = Attendance.objects.filter(session__session_date=today).count()

    # ── Chart data ──
    # Donut: student tajweed levels (using gender as proxy since no tajweed_level field)
    student_levels = {
        "excellent": User.objects.filter(role=User.Role.STUDENT, is_approved=User.ApprovalStatus.APPROVED, gender="female").count(),
        "good": User.objects.filter(role=User.Role.STUDENT, is_approved=User.ApprovalStatus.APPROVED, gender="male").count(),
        "weak": User.objects.filter(role=User.Role.STUDENT, is_approved=User.ApprovalStatus.APPROVED).exclude(gender__in=["male", "female"]).count(),
    }

    # Line chart: last 7 months attendance trend
    months_data = []
    now = timezone.now()
    from django.utils.dateparse import parse_date
    for i in range(6, -1, -1):
        month_start = now.replace(day=1) - timedelta(days=30 * i)
        month_end = (month_start.replace(day=28) + timedelta(days=4)).replace(day=1) - timedelta(days=1)
        present = Attendance.objects.filter(
            session__session_date__gte=month_start.date(),
            session__session_date__lte=month_end.date(),
            status__in=["present", "late"]
        ).count()
        absent = Attendance.objects.filter(
            session__session_date__gte=month_start.date(),
            session__session_date__lte=month_end.date(),
            status="absent"
        ).count()
        months_data.append({
            "month": month_start.strftime("%b"),
            "present": present,
            "absent": absent,
        })

    recent_certificates = Certificate.objects.select_related("student", "template", "issued_by").order_by("-issue_date")[:5]

    # Pending circle enrollment requests
    pending_enrollments = CircleEnrollment.objects.filter(
        status=CircleEnrollment.Status.PENDING
    ).select_related('student', 'circle', 'circle__teacher').order_by('enrolled_at')[:5]
    pending_enrollment_count = CircleEnrollment.objects.filter(
        status=CircleEnrollment.Status.PENDING
    ).count()

    # Repeated absentees (absent >3 times this month)
    this_month_start = now.replace(day=1).date()
    repeated_absentees = Attendance.objects.filter(
        status="absent", session__session_date__gte=this_month_start
    ).values("student_id").annotate(cnt=Count("id")).filter(cnt__gt=3).count()

    # Inactive circles (no session in 3 weeks)
    three_weeks_ago = date.today() - timedelta(weeks=3)
    active_circle_ids = Session.objects.filter(
        session_date__gte=three_weeks_ago
    ).values_list("circle_id", flat=True).distinct()
    inactive_circles = Circle.objects.filter(
        status=Circle.Status.ACTIVE
    ).exclude(id__in=active_circle_ids).count()

    return render(request, 'dashboard/home.html', {
        'stats': {
            'total_students': total_students,
            'active_circles': active_circles,
            'total_teachers': total_teachers,
            'total_supervisors': total_supervisors,
            'attendance_rate': attendance_rate,
            'total_absent': total_absent,
            'weekly_attendance': weekly_attendance,
            'circle_distribution': circle_distribution,
            'recent_users': recent_users,
            'pending_users': pending_users,
            'pending_count': pending_count,
            'total_hifz_thumns': total_hifz_thumns,
            'total_hifz_units': format_hizb_thumn(total_hifz_thumns),
            'total_murajaa_thumns': total_murajaa_thumns,
            'total_murajaa_units': format_hizb_thumn(total_murajaa_thumns),
            'total_certificates': total_certificates,
        },
        'pending_count': pending_count,
        'approved_count': approved_count,
        'rejected_count': rejected_count,
        'recent_requests': recent_requests,
        'recent_certificates': recent_certificates,
        'pending_enrollments': pending_enrollments,
        'pending_enrollment_count': pending_enrollment_count,
        'urgent_requests': urgent_requests,
        'today_sessions': today_sessions,
        'today_attendance': today_attendance,
        'chart_donut': json.dumps(student_levels),
        'chart_line': json.dumps(months_data),
        'repeated_absentees': repeated_absentees,
        'inactive_circles': inactive_circles,
    })
@login_required
@role_required(User.Role.MAIN_ADMIN, User.Role.SUB_ADMIN)
def admin_inscriptions(request):

    batch_ids = _scoped_batch_ids(request.user)
    status = request.GET.get('status', 'pending')
    role_filter = request.GET.get('role', '')
    search = request.GET.get('search', '')
    role_tab = request.GET.get('role_tab', '')

    users = User.objects.all().order_by('-created_at')
    if batch_ids is not None:
        # Pending signups have no batch yet — keep them visible so the
        # sub-admin can approve them (batch gets set on approval).
        users = users.filter(
            Q(batch_id__in=batch_ids) | Q(batch__isnull=True, is_approved='pending')
        )

    if status == 'approved':
        users = users.filter(is_approved='approved')
    elif status == 'rejected':
        users = users.filter(is_approved='rejected')
    else:
        users = users.filter(is_approved='pending')

    if role_filter:
        users = users.filter(role=role_filter)
    if search:
        users = users.filter(Q(full_name_ar__icontains=search) | Q(email__icontains=search))

    paginator = Paginator(users, 15)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)

    user_qs = User.objects.all()
    if batch_ids is not None:
        user_qs = user_qs.filter(
            Q(batch_id__in=batch_ids) | Q(batch__isnull=True, is_approved='pending')
        )

    pending_count = user_qs.filter(is_approved='pending').count()
    approved_count = user_qs.filter(is_approved='approved').count()
    rejected_count = user_qs.filter(is_approved='rejected').count()

    # Role-based grouped data for the role tab view
    supervisor_requests = user_qs.filter(is_approved='pending', role=User.Role.SUB_ADMIN)
    teacher_requests = user_qs.filter(is_approved='pending', role=User.Role.TEACHER)
    student_requests = user_qs.filter(is_approved='pending', role=User.Role.STUDENT)

    # Group student requests by enrollment circle
    student_groups = {}
    for s in student_requests.prefetch_related('enrollments__circle'):
        circle_name = s.enrollments.filter(status=CircleEnrollment.Status.ACTIVE).first()
        group_key = circle_name.circle.name if circle_name else "غير مسجل"
        student_groups.setdefault(group_key, []).append(s)

    # Stats for the top row
    circle_qs = Circle.objects.filter(status=Circle.Status.ACTIVE)
    if batch_ids is not None:
        circle_qs = circle_qs.filter(batch_id__in=batch_ids)
        total_teachers_count = user_qs.filter(role=User.Role.TEACHER, is_approved=User.ApprovalStatus.APPROVED).count()
        total_students_count = user_qs.filter(role=User.Role.STUDENT, is_approved=User.ApprovalStatus.APPROVED).count()
    else:
        total_teachers_count = User.objects.filter(role=User.Role.TEACHER, is_approved=User.ApprovalStatus.APPROVED).count()
        total_students_count = User.objects.filter(role=User.Role.STUDENT, is_approved=User.ApprovalStatus.APPROVED).count()
    total_circles = circle_qs.count()

    return render(request, 'dashboard/inscriptions.html', {
        'active_batches': Batch.objects.filter(status=Batch.Status.ACTIVE),
        'pending_users': page_obj,
        'pending_count': pending_count,
        'approved_count': approved_count,
        'rejected_count': rejected_count,
        'supervisor_requests': supervisor_requests,
        'teacher_requests': teacher_requests,
        'student_requests_by_group': student_groups,
        'total_circles': total_circles,
        'total_teachers_count': total_teachers_count,
        'total_students_count': total_students_count,
        'role_tab': role_tab,
    })
@login_required
@role_required(User.Role.MAIN_ADMIN, User.Role.SUB_ADMIN)
def approve_user(request, pk):

    user_obj = get_object_or_404(User, pk=pk)

    # A Sub Admin may only act on users visible in their scope (Rule: no
    # cross-batch actions). Pending users without a batch stay actionable.
    if scoping.scoped_pending_users(
        request.user, User.objects.filter(pk=user_obj.pk)
    ).first() is None:
        raise PermissionDenied

    toast_msg = ""
    if request.method == "POST":
        form = ApprovalForm(request.POST)
        if form.is_valid():
            action = form.cleaned_data["action"]
            if action == "approve":
                user_obj.is_approved = User.ApprovalStatus.APPROVED
                user_obj.rejection_reason = ""
                # الدفعة: sub-admins place the user in their own batch;
                # the main admin may pick one in the approve form.
                if user_obj.batch_id is None:
                    if request.user.role == User.Role.SUB_ADMIN:
                        user_obj.batch = scoping.scoped_batch(request.user)
                    else:
                        user_obj.batch = form.cleaned_data.get("batch")
                user_obj.save(update_fields=[
                    "is_approved", "rejection_reason", "batch", "updated_at",
                ])
                toast_msg = "تم اعتماد المستخدم بنجاح"
                email_sent = send_approval_email(
                    user=user_obj,
                    login_url=request.build_absolute_uri(reverse("accounts:login")),
                )
                if not email_sent:
                    logger.error("Failed to send approval email to %s", user_obj.email)
            elif action == "reject":
                reason = form.cleaned_data.get("rejection_reason", "")
                user_obj.is_approved = User.ApprovalStatus.REJECTED
                user_obj.rejection_reason = reason
                user_obj.save(update_fields=["is_approved", "rejection_reason", "updated_at"])
                toast_msg = "تم رفض المستخدم"
                email_sent = send_rejection_email(
                    user=user_obj,
                    reason=reason,
                )
                if not email_sent:
                    logger.error("Failed to send rejection email to %s", user_obj.email)

    response = render(request, "dashboard/partials/user_row.html", {
        "user_obj": user_obj,
        "active_batches": Batch.objects.filter(status=Batch.Status.ACTIVE),
    })
    if toast_msg:
        response["HX-Trigger"] = json.dumps({"showToast": {"message": toast_msg, "type": "success"}})
    return response
@login_required
@role_required(User.Role.MAIN_ADMIN, User.Role.SUB_ADMIN)
def pending_users_table(request):

    filter_status = request.GET.get("status", "pending")
    role = request.GET.get("role", "")
    search = request.GET.get("search", "")

    if filter_status == "approved":
        users = scoping.scoped_users(
            request.user, User.objects.filter(is_approved=User.ApprovalStatus.APPROVED)
        )
    elif filter_status == "rejected":
        users = scoping.scoped_pending_users(
            request.user, User.objects.filter(is_approved=User.ApprovalStatus.REJECTED)
        )
    else:
        users = scoping.scoped_pending_users(
            request.user, User.objects.filter(is_approved=User.ApprovalStatus.PENDING)
        )

    if role:
        users = users.filter(role=role)
    if search:
        users = users.filter(Q(full_name_ar__icontains=search) | Q(email__icontains=search))

    users = users.order_by("-created_at")
    return render(request, "dashboard/partials/user_rows.html", {
        "users": users,
        "active_batches": Batch.objects.filter(status=Batch.Status.ACTIVE),
    })
@login_required
@role_required(User.Role.MAIN_ADMIN, User.Role.SUB_ADMIN)
def admin_teacher_absences(request):

    status_filter = request.GET.get("status", "")
    qs = TeacherAbsence.objects.select_related("teacher", "substitute_teacher").order_by("-created_at")
    if status_filter:
        qs = qs.filter(status=status_filter)

    pending_count = TeacherAbsence.objects.filter(status=TeacherAbsence.Status.PENDING).count()
    approved_count = TeacherAbsence.objects.filter(status=TeacherAbsence.Status.APPROVED).count()
    rejected_count = TeacherAbsence.objects.filter(status=TeacherAbsence.Status.REJECTED).count()

    paginator = Paginator(qs, 15)
    page_obj = paginator.get_page(request.GET.get("page", 1))

    return render(request, "dashboard/absences/list.html", {
        "absences": page_obj,
        "page_obj": page_obj,
        "pending_count": pending_count,
        "approved_count": approved_count,
        "rejected_count": rejected_count,
    })
@login_required
@role_required(User.Role.MAIN_ADMIN, User.Role.SUB_ADMIN)
def admin_absence_manage(request, pk):
    absence = get_object_or_404(
        TeacherAbsence.objects.select_related("teacher", "substitute_teacher", "processed_by"),
        pk=pk,
    )
    teachers = User.objects.filter(
        role=User.Role.TEACHER, is_approved=User.ApprovalStatus.APPROVED, is_active=True,
    ).exclude(pk=absence.teacher.pk)

    circles = absence.teacher.teaching_circles.filter(status=Circle.Status.ACTIVE)
    substitutions = TeacherSubstitution.objects.filter(absence=absence).select_related("circle", "substitute_teacher")

    if request.method == "POST":
        action = request.POST.get("action", "")
        if action == "approve":
            absence.status = TeacherAbsence.Status.APPROVED
            absence.processed_by = request.user
            absence.save()

            # Create per-circle substitutions
            for c in circles:
                sub_teacher_pk = request.POST.get(f"sub_{c.pk}", "")
                TeacherSubstitution.objects.update_or_create(
                    absence=absence, circle=c,
                    defaults={"substitute_teacher_id": sub_teacher_pk or None},
                )

            messages.success(request, "تم قبول طلب الغياب وتعيين البديل للحلقات")
            return redirect("accounts:admin_teacher_absences")
        elif action == "reject":
            absence.status = TeacherAbsence.Status.REJECTED
            absence.rejection_reason = request.POST.get("rejection_reason", "")
            absence.processed_by = request.user
            absence.save()
            messages.success(request, "تم رفض طلب الغياب")
            return redirect("accounts:admin_teacher_absences")

    sub_map = {s.circle_id: s for s in substitutions}
    circles_with_sub = [(c, sub_map.get(c.id)) for c in circles]

    return render(request, "dashboard/absences/manage.html", {
        "absence": absence,
        "teachers": teachers,
        "circles": circles_with_sub,
    })
@login_required
@role_required(User.Role.MAIN_ADMIN, User.Role.SUB_ADMIN)
def admin_active_substitutions(request):

    today = date.today()
    active_absences = TeacherAbsence.objects.filter(
        status=TeacherAbsence.Status.APPROVED,
        start_date__lte=today, end_date__gte=today,
    ).select_related("teacher", "substitute_teacher", "processed_by").order_by("start_date")

    substitutions = TeacherSubstitution.objects.filter(
        absence__in=active_absences,
    ).select_related("circle", "substitute_teacher", "absence__teacher")

    return render(request, "dashboard/absences/active.html", {
        "absences": active_absences,
        "substitutions": substitutions,
    })
@login_required
@role_required(User.Role.MAIN_ADMIN, User.Role.SUB_ADMIN)
def admin_teachers(request):

    batch_ids = scoping.scoped_batch_ids(request.user)
    search = request.GET.get("search", "")
    status_filter = request.GET.get("status", "")

    base_qs = User.objects.filter(role=User.Role.TEACHER, is_approved=User.ApprovalStatus.APPROVED)
    if batch_ids is not None:
        base_qs = base_qs.filter(batch_id__in=batch_ids)
    teachers = base_qs.annotate(
        circles_count=Count('teaching_circles', filter=Q(teaching_circles__status=Circle.Status.ACTIVE)),
        students_count=Count('teaching_circles__enrollments', filter=Q(teaching_circles__enrollments__status=CircleEnrollment.Status.ACTIVE)),
        sessions_count=Count('teaching_circles__sessions'),
    ).order_by('-created_at')

    if status_filter == "active":
        teachers = teachers.filter(is_active=True)
    elif status_filter == "inactive":
        teachers = teachers.filter(is_active=False)

    if search:
        teachers = teachers.filter(Q(full_name_ar__icontains=search) | Q(email__icontains=search))

    total_teachers = base_qs.count()
    active_teachers = base_qs.filter(is_active=True).count()
    inactive_teachers = total_teachers - active_teachers
    active_circles_count = Circle.objects.filter(
        teacher__in=base_qs, status=Circle.Status.ACTIVE
    ).count()

    paginator = Paginator(teachers, 15)
    page_obj = paginator.get_page(request.GET.get('page', 1))

    circles = Circle.objects.filter(status=Circle.Status.ACTIVE).order_by('name')

    return render(request, 'dashboard/teachers/list.html', {
        'teachers': page_obj,
        'page_obj': page_obj,
        'total_teachers': total_teachers,
        'active_teachers': active_teachers,
        'inactive_teachers': inactive_teachers,
        'active_circles_count': active_circles_count,
        'circles': circles,
    })
@login_required
@role_required(User.Role.MAIN_ADMIN, User.Role.SUB_ADMIN)
def admin_students(request):

    batch_ids = scoping.scoped_batch_ids(request.user)
    search = request.GET.get("search", "")
    status_filter = request.GET.get("status", "")
    circle_filter = request.GET.get("circle", "")

    base_qs = User.objects.filter(role=User.Role.STUDENT, is_approved=User.ApprovalStatus.APPROVED)
    if batch_ids is not None:
        base_qs = base_qs.filter(batch_id__in=batch_ids)
    students = base_qs.annotate(
        active_enrollments=Count('enrollments', filter=Q(enrollments__status=CircleEnrollment.Status.ACTIVE)),
        total_sessions=Count('attendance_records', distinct=True),
        present_sessions=Count('attendance_records', filter=Q(attendance_records__status__in=['present', 'late', 'left_early']), distinct=True),
    ).order_by('-created_at')

    if status_filter == "active":
        students = students.filter(enrollments__status=CircleEnrollment.Status.ACTIVE).distinct()
    elif status_filter == "inactive":
        students = students.exclude(enrollments__status=CircleEnrollment.Status.ACTIVE).distinct()

    if circle_filter:
        students = students.filter(enrollments__circle_id=circle_filter, enrollments__status=CircleEnrollment.Status.ACTIVE)

    if search:
        students = students.filter(Q(full_name_ar__icontains=search) | Q(email__icontains=search) | Q(phone__icontains=search))

    total_students = base_qs.count()
    active_base = User.objects.filter(
        role=User.Role.STUDENT, enrollments__status=CircleEnrollment.Status.ACTIVE,
    )
    if batch_ids is not None:
        active_base = active_base.filter(batch_id__in=batch_ids)
    active_students = active_base.distinct().count()
    inactive_students = total_students - active_students

    circles = Circle.objects.filter(status=Circle.Status.ACTIVE)
    if batch_ids is not None:
        circles = circles.filter(batch_id__in=batch_ids)

    paginator = Paginator(students, 15)
    page_obj = paginator.get_page(request.GET.get('page', 1))

    return render(request, 'dashboard/students/list.html', {
        'students': page_obj,
        'page_obj': page_obj,
        'total_students': total_students,
        'active_students': active_students,
        'inactive_students': inactive_students,
        'circles': circles,
    })
@login_required
@role_required(User.Role.MAIN_ADMIN, User.Role.SUB_ADMIN)
def admin_student_detail(request, pk):

    student = get_object_or_404(User, pk=pk, role=User.Role.STUDENT)
    _check_batch_access(request.user, student.batch_id)

    enrollments = CircleEnrollment.objects.filter(student=student).select_related(
        "circle__teacher"
    ).order_by("-enrolled_at")

    active_enrollments = enrollments.filter(status=CircleEnrollment.Status.ACTIVE)

    total_sessions = Session.objects.filter(
        attendance_records__student=student
    ).count()
    present_count = Attendance.objects.filter(
        student=student, status__in=["present", "late"]
    ).count()
    attendance_rate = round(present_count / total_sessions * 100) if total_sessions else 0

    hifz_total = count_thumns(
        MemorizationProgress.objects.filter(
            enrollment__student=student, type="hifz"
        ).values_list("surah_id", "ayah_from", "ayah_to")
    )
    mastered = count_thumns(
        MemorizationProgress.objects.filter(
            enrollment__student=student, type="hifz", status="mastered"
        ).values_list("surah_id", "ayah_from", "ayah_to")
    )

    recent_attendance = Attendance.objects.filter(student=student).select_related(
        "session__circle"
    ).order_by("-session__session_date")[:20]

    absence_breakdown = Attendance.objects.filter(
        student=student, status__in=[Attendance.Status.ABSENT, Attendance.Status.ABSENT_UNJUSTIFIED, Attendance.Status.ABSENT_JUSTIFIED]
    ).aggregate(
        total=Count('id'),
        justified=Count('id', filter=Q(justification_status=Attendance.JustificationStatus.ACCEPTED)),
        pending=Count('id', filter=Q(justification_status=Attendance.JustificationStatus.PENDING)),
        unjustified=Count('id', filter=Q(justification_status__in=[
            Attendance.JustificationStatus.REFUSED, Attendance.JustificationStatus.NONE,
        ])),
    )


    return render(request, "dashboard/students/detail.html", {
        "student": student,
        "enrollments": enrollments,
        "active_enrollments": active_enrollments,
        "total_sessions": total_sessions,
        "attendance_rate": attendance_rate,
        "present_count": present_count,
        "absence_breakdown": absence_breakdown,
        "hifz_total": hifz_total,
        "hifz_units": format_hizb_thumn(hifz_total),
        "mastered": mastered,
        "mastered_units": format_hizb_thumn(mastered),
        "recent_attendance": recent_attendance,
    })
@login_required
@role_required(User.Role.MAIN_ADMIN, User.Role.SUB_ADMIN)
def admin_student_toggle_status(request, pk):
    student = get_object_or_404(User, pk=pk, role=User.Role.STUDENT)
    student.is_active = not student.is_active
    student.save(update_fields=["is_active"])
    return redirect("accounts:admin_student_detail", pk=pk)
@login_required
@role_required(User.Role.MAIN_ADMIN, User.Role.SUB_ADMIN)
def admin_student_create(request):

    if request.method == "POST":
        form = SignupForm(request.POST)
        if form.is_valid():
            user = form.save()
            user.role = User.Role.STUDENT
            user.is_approved = User.ApprovalStatus.APPROVED
            user.save()
            messages.success(request, f"تم إضافة الطالب {user.full_name_ar} بنجاح")
            return redirect("accounts:admin_students")
    else:
        form = SignupForm()

    return render(request, "dashboard/students/create.html", {"form": form})


@login_required
@role_required(User.Role.MAIN_ADMIN, User.Role.SUB_ADMIN)
def admin_student_edit(request, pk):
    student = get_object_or_404(User, pk=pk, role=User.Role.STUDENT)
    _check_batch_access(request.user, student.batch_id)

    if request.method == "POST":
        form = ProfileForm(request.POST, instance=student)
        if form.is_valid():
            student = form.save(commit=False)
            student.role = User.Role.STUDENT
            student.is_active = request.POST.get("is_active") == "true"
            if request.user.role == User.Role.MAIN_ADMIN:
                batch_id = request.POST.get("batch") or None
                if batch_id is None or Batch.objects.filter(pk=batch_id).exists():
                    student.batch_id = batch_id
            student.save()
            messages.success(request, "تم تحديث بيانات الطالب بنجاح")
            return redirect("accounts:admin_student_detail", pk=student.pk)
    else:
        form = ProfileForm(instance=student)

    return render(request, "dashboard/students/edit.html", {
        "form": form,
        "student": student,
        "active_batches": Batch.objects.filter(status=Batch.Status.ACTIVE),
    })


# ─── Export: PDF for list pages ──────────────────────

# ── Branded PDF helpers (shared by the list + report exports) ──────────
# Palette mirrors the web design system (emerald/teal + navy + gold).
_PDF_TEAL = (42, 168, 168)
_PDF_DARK = (26, 60, 78)
_PDF_GOLD = (201, 168, 76)
_PDF_GRAY = (107, 138, 154)
_PDF_LIGHT = (236, 246, 246)
_PDF_LINE = (214, 230, 230)
_PDF_WHITE = (255, 255, 255)

_NOTO_DIR = "/usr/share/fonts/truetype/noto/"
_LOGO_ASPECT = 500 / 646  # width / height of static/img/logo.png


def _pdf_logo_path():
    from django.conf import settings
    p = settings.BASE_DIR / "static" / "img" / "logo.png"
    return str(p) if p.exists() else None


def _pdf_ltr(val):
    """Wrap a purely non-Arabic value (phone, email, id, url) in a Unicode
    left-to-right isolate so the RTL table base direction doesn't displace a
    leading '+' or reorder mixed digits. The isolate chars are default-
    ignorable, so they add no visible glyph."""
    import re
    if val and val != "—" and not re.search(r"[؀-ۿ]", val):
        return f"⁦{val}⁩"  # LRI … PDI
    return val


def _make_branded_pdf(orientation, title, subtitle=""):
    """A Hafez-branded FPDF: full masthead (logo + institution + title bar) on
    page 1, a slim running header on later pages, and a branded footer."""
    from fpdf import FPDF

    class BrandedPDF(FPDF):
        doc_title = title
        doc_subtitle = subtitle
        _masthead_done = False

        def header(self):
            logo = _pdf_logo_path()
            if not self._masthead_done:
                self._draw_masthead(logo)
                self._masthead_done = True
            else:
                self._draw_running_header(logo)

        def _draw_masthead(self, logo):
            lh = 20.0
            lw = lh * _LOGO_ASPECT
            if logo:
                self.image(logo, x=self.w - self.r_margin - lw, y=10, h=lh)
            text_w = (self.w - self.l_margin - self.r_margin) - lw - 4
            self.set_xy(self.l_margin, 12)
            self.set_font("Arabic", "B", 15)
            self.set_text_color(*_PDF_DARK)
            self.cell(text_w, 8, "الطبيب الحافظ", align="R")
            self.set_xy(self.l_margin, 21)
            self.set_font("Arabic", "", 8.5)
            self.set_text_color(*_PDF_GRAY)
            self.cell(text_w, 5, "كلية الطب - قسنطينة   •   منصة تحفيظ القرآن", align="R")
            # gold hairline under the masthead
            self.set_draw_color(*_PDF_GOLD)
            self.set_line_width(0.6)
            self.line(self.l_margin, 33, self.w - self.r_margin, 33)
            # teal title bar
            self._title_bar(37)

        def _title_bar(self, y):
            h = 12.0
            w = self.w - self.l_margin - self.r_margin
            self.set_fill_color(*_PDF_TEAL)
            self.set_draw_color(*_PDF_TEAL)
            try:
                self.rect(self.l_margin, y, w, h, style="F",
                          round_corners=True, corner_radius=2)
            except TypeError:
                self.rect(self.l_margin, y, w, h, style="F")
            self.set_xy(self.l_margin, y)
            self.set_font("Arabic", "B", 14)
            self.set_text_color(*_PDF_WHITE)
            self.cell(w, h, self.doc_title, align="C")
            self.set_xy(self.l_margin, y + h + 2.5)
            if self.doc_subtitle:
                self.set_font("Arabic", "", 9)
                self.set_text_color(*_PDF_GRAY)
                self.cell(w, 6, self.doc_subtitle, align="C")
                self.set_y(y + h + 11)
            else:
                self.set_y(y + h + 5)

        def _draw_running_header(self, logo):
            if logo:
                lh = 8.0
                self.image(logo, x=self.w - self.r_margin - lh * _LOGO_ASPECT, y=7, h=lh)
            self.set_xy(self.l_margin, 8)
            self.set_font("Arabic", "B", 9)
            self.set_text_color(*_PDF_DARK)
            self.cell(0, 8, self.doc_title, align="R")
            self.set_draw_color(*_PDF_TEAL)
            self.set_line_width(0.4)
            self.line(self.l_margin, 17, self.w - self.r_margin, 17)
            self.set_y(21)

        def footer(self):
            self.set_y(-15)
            self.set_draw_color(*_PDF_LINE)
            self.set_line_width(0.3)
            self.line(self.l_margin, self.get_y(), self.w - self.r_margin, self.get_y())
            self.set_xy(self.l_margin, -12)
            self.set_font("Arabic", "", 7.5)
            self.set_text_color(*_PDF_GRAY)
            self.cell(0, 6, "منصة الطبيب الحافظ", align="R")
            self.set_xy(self.l_margin, -12)
            self.cell(0, 6, f"تاريخ التصدير: {date.today().strftime('%Y-%m-%d')}", align="L")
            # Page x/y: draw unshaped so the {nb} total-pages alias survives to
            # output (RTL shaping turns text into glyphs and breaks the alias).
            saved_shaping = self.text_shaping
            self.text_shaping = None
            self.set_xy(self.l_margin, -12)
            self.set_font("DejaVu", "", 7.5)
            self.cell(0, 6, f"{self.page_no()} / {{nb}}", align="C")
            self.text_shaping = saved_shaping

    pdf = BrandedPDF(orientation=orientation, unit="mm", format="A4")
    pdf.set_margins(10, 10, 10)
    pdf.set_auto_page_break(auto=True, margin=20)
    pdf.add_font("Arabic", "", _NOTO_DIR + "NotoSansArabic-Regular.ttf")
    pdf.add_font("Arabic", "B", _NOTO_DIR + "NotoSansArabic-Bold.ttf")
    pdf.add_font("DejaVu", "", "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf")
    pdf.set_fallback_fonts(["DejaVu"])
    pdf.alias_nb_pages()
    # Proper Arabic shaping + Unicode bidi: RTL is the base paragraph
    # direction, but embedded number/Latin runs (phones, emails, %, counts)
    # keep their natural left-to-right order instead of being reversed.
    pdf.set_text_shaping(True, direction="rtl")
    return pdf


def _export_list_pdf(request, title, headers, rows, filename):
    from io import BytesIO

    landscape = len(headers) > 5
    subtitle = (f"تاريخ التصدير: {date.today().strftime('%Y-%m-%d')}"
                f"      •      إجمالي السجلات: {len(rows)}")
    pdf = _make_branded_pdf("L" if landscape else "P", title, subtitle)
    pdf.add_page()

    # Column widths — proportional to header length, normalized to full width.
    avail = pdf.w - pdf.l_margin - pdf.r_margin
    if headers:
        weights = [max(len(h), 8) for h in headers]
        col_w = [avail * w / sum(weights) for w in weights]
    else:
        col_w = [avail]
    head_h = 10.0
    row_h = 8.5

    def draw_header_row():
        pdf.set_fill_color(*_PDF_DARK)
        pdf.set_text_color(*_PDF_WHITE)
        pdf.set_draw_color(*_PDF_DARK)
        pdf.set_font("Arabic", "B", 9)
        for i, h in enumerate(headers):
            pdf.cell(col_w[i], head_h, h, border=0, align="C", fill=True)
        pdf.ln()

    draw_header_row()

    if not rows:
        pdf.set_font("Arabic", "", 9)
        pdf.set_text_color(*_PDF_GRAY)
        pdf.set_fill_color(*_PDF_LIGHT)
        pdf.cell(sum(col_w), 12, "لا توجد بيانات للعرض", border=0, align="C", fill=True)
        pdf.ln()

    pdf.set_font("Arabic", "", 8)
    for idx, row in enumerate(rows):
        # Repeat the column header after an automatic page break.
        if pdf.will_page_break(row_h):
            pdf.add_page()
            draw_header_row()
            pdf.set_font("Arabic", "", 8)
        pdf.set_fill_color(*(_PDF_LIGHT if idx % 2 == 0 else _PDF_WHITE))
        pdf.set_text_color(*_PDF_DARK)
        pdf.set_draw_color(*_PDF_LINE)
        for i, cell_val in enumerate(row):
            val = str(cell_val) if cell_val not in (None, "") else "—"
            pdf.cell(col_w[i], row_h, _pdf_ltr(val), border="B", align="C", fill=True)
        pdf.ln()

    # Total chip
    pdf.ln(6)
    chip_w = 70.0
    chip_x = pdf.l_margin + (avail - chip_w) / 2
    pdf.set_xy(chip_x, pdf.get_y())
    pdf.set_fill_color(*_PDF_LIGHT)
    pdf.set_text_color(*_PDF_DARK)
    pdf.set_font("Arabic", "B", 10)
    try:
        pdf.cell(chip_w, 10, f"إجمالي السجلات: {len(rows)}", align="C", fill=True,
                 border=0, round_corners=True)
    except TypeError:
        pdf.cell(chip_w, 10, f"إجمالي السجلات: {len(rows)}", align="C", fill=True)

    buf = BytesIO()
    pdf.output(buf)
    buf.seek(0)
    response = HttpResponse(buf.read(), content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{filename}.pdf"'
    return response


def _export_list_excel(request, title, headers, rows, filename):
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title[:31]

    # Colors matching the design system
    TEAL = "2AA8A8"
    DARK = "1A3C4E"
    LIGHT_BG = "EBF5F5"
    WHITE = "FFFFFF"
    GOLD = "C9A84C"

    thin_border = Border(
        left=Side(style='thin', color='D0E8E8'),
        right=Side(style='thin', color='D0E8E8'),
        top=Side(style='thin', color='D0E8E8'),
        bottom=Side(style='thin', color='D0E8E8'),
    )

    header_fill = PatternFill(start_color=TEAL, end_color=TEAL, fill_type="solid")
    header_font = Font(name="Noto Sans Arabic", size=11, bold=True, color=WHITE)
    cell_font = Font(name="Noto Sans Arabic", size=10)
    title_font = Font(name="Noto Sans Arabic", size=16, bold=True, color=DARK)
    alt_fill = PatternFill(start_color=LIGHT_BG, end_color=LIGHT_BG, fill_type="solid")

    # Title row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.font = title_font
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # Subtitle row
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    sub_cell = ws.cell(row=2, column=1, value=f"تاريخ التصدير: {date.today().strftime('%Y-%m-%d')} | إجمالي السجلات: {len(rows)}")
    sub_cell.font = Font(name="Noto Sans Arabic", size=9, color="6B8A9A")
    sub_cell.alignment = Alignment(horizontal="center")

    # Header row
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=i, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
    ws.row_dimensions[4].height = 25

    # Data rows
    for r_idx, row in enumerate(rows, 5):
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.font = cell_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
            if (r_idx - 5) % 2 == 1:
                cell.fill = alt_fill

    # Summary row
    summary_row = 5 + len(rows)
    ws.merge_cells(start_row=summary_row, start_column=1, end_row=summary_row, end_column=len(headers))
    summary_cell = ws.cell(row=summary_row, column=1, value=f"إجمالي السجلات: {len(rows)}")
    summary_cell.font = Font(name="Noto Sans Arabic", size=11, bold=True, color=TEAL)
    summary_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[summary_row].height = 28

    # Auto-fit column widths
    for col_cells in ws.columns:
        max_len = 0
        for cell in col_cells:
            if isinstance(cell, openpyxl.cell.cell.MergedCell):
                continue
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        first_cell = col_cells[0]
        if not isinstance(first_cell, openpyxl.cell.cell.MergedCell):
            ws.column_dimensions[first_cell.column_letter].width = min(max_len + 4, 45)

    # Freeze header row
    ws.freeze_panes = "A5"

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}.xlsx"'
    wb.save(response)
    return response
@login_required
@role_required(User.Role.MAIN_ADMIN, User.Role.SUB_ADMIN)
def admin_students_export_pdf(request):
    students = User.objects.filter(role=User.Role.STUDENT, is_approved=User.ApprovalStatus.APPROVED)
    headers = ["الاسم", "البريد", "الهاتف", "الجنس", "الحلقات", "الحالة"]
    rows = [[s.full_name_ar, s.email, s.phone, {"male": "ذكر", "female": "أنثى"}.get(s.gender, s.gender) if s.gender else "—",
             s.enrollments.filter(status=CircleEnrollment.Status.ACTIVE).count(),
             "نشط" if s.is_active else "غير نشط"] for s in students]
    return _export_list_pdf(request, "قائمة الطلاب", headers, rows, "students_list")
@login_required
@role_required(User.Role.MAIN_ADMIN, User.Role.SUB_ADMIN)
def admin_students_export_excel(request):
    students = User.objects.filter(role=User.Role.STUDENT, is_approved=User.ApprovalStatus.APPROVED)
    headers = ["الاسم", "البريد", "الهاتف", "الجنس", "الحلقات", "الحالة"]
    rows = [[s.full_name_ar, s.email, s.phone, {"male": "ذكر", "female": "أنثى"}.get(s.gender, s.gender) if s.gender else "—",
             s.enrollments.filter(status=CircleEnrollment.Status.ACTIVE).count(),
             "نشط" if s.is_active else "غير نشط"] for s in students]
    return _export_list_excel(request, "قائمة الطلاب", headers, rows, "students_list")
@login_required
@role_required(User.Role.MAIN_ADMIN, User.Role.SUB_ADMIN)
def admin_teachers_export_pdf(request):
    teachers = User.objects.filter(role=User.Role.TEACHER, is_approved=User.ApprovalStatus.APPROVED).annotate(
        circles_count=Count('teaching_circles', filter=Q(teaching_circles__status=Circle.Status.ACTIVE)),
        students_count=Count('teaching_circles__enrollments', filter=Q(teaching_circles__enrollments__status=CircleEnrollment.Status.ACTIVE)),
    )
    headers = ["الاسم", "البريد", "الهاتف", "الحلقات", "الطلاب", "الحالة"]
    rows = [[t.full_name_ar, t.email, t.phone, t.circles_count, t.students_count,
             "نشط" if t.is_active else "غير نشط"] for t in teachers]
    return _export_list_pdf(request, "قائمة الأساتذة", headers, rows, "teachers_list")
@login_required
@role_required(User.Role.MAIN_ADMIN, User.Role.SUB_ADMIN)
def admin_teachers_export_excel(request):
    teachers = User.objects.filter(role=User.Role.TEACHER, is_approved=User.ApprovalStatus.APPROVED).annotate(
        circles_count=Count('teaching_circles', filter=Q(teaching_circles__status=Circle.Status.ACTIVE)),
        students_count=Count('teaching_circles__enrollments', filter=Q(teaching_circles__enrollments__status=CircleEnrollment.Status.ACTIVE)),
    )
    headers = ["الاسم", "البريد", "الهاتف", "الحلقات", "الطلاب", "الحالة"]
    rows = [[t.full_name_ar, t.email, t.phone, t.circles_count, t.students_count,
             "نشط" if t.is_active else "غير نشط"] for t in teachers]
    return _export_list_excel(request, "قائمة الأساتذة", headers, rows, "teachers_list")
@login_required
@role_required(User.Role.MAIN_ADMIN, User.Role.SUB_ADMIN)
def admin_circles_export_pdf(request):
    circles = Circle.objects.select_related("teacher").annotate(
        student_count=Count('enrollments', filter=Q(enrollments__status=CircleEnrollment.Status.ACTIVE)),
        sessions_count=Count('sessions'),
    )
    headers = ["اسم الحلقة", "المعلم", "الجنس", "الطلاب", "الحصص", "الحالة"]
    rows = [[c.name, c.teacher.full_name_ar if c.teacher else "—",
             "ذكور" if c.gender == "male" else "إناث",
             c.student_count, c.sessions_count,
             c.get_status_display()] for c in circles]
    return _export_list_pdf(request, "قائمة الحلقات", headers, rows, "circles_list")
@login_required
@role_required(User.Role.MAIN_ADMIN, User.Role.SUB_ADMIN)
def admin_circles_export_excel(request):
    circles = Circle.objects.select_related("teacher").annotate(
        student_count=Count('enrollments', filter=Q(enrollments__status=CircleEnrollment.Status.ACTIVE)),
        sessions_count=Count('sessions'),
    )
    headers = ["اسم الحلقة", "المعلم", "الجنس", "الطلاب", "الحصص", "الحالة"]
    rows = [[c.name, c.teacher.full_name_ar if c.teacher else "—",
             "ذكور" if c.gender == "male" else "إناث",
             c.student_count, c.sessions_count,
             c.get_status_display()] for c in circles]
    return _export_list_excel(request, "قائمة الحلقات", headers, rows, "circles_list")
@login_required
@role_required(User.Role.MAIN_ADMIN, User.Role.SUB_ADMIN)
def admin_inscriptions_export_pdf(request):
    users = User.objects.filter(is_approved=User.ApprovalStatus.PENDING)
    headers = ["الاسم", "البريد", "الدور", "الهاتف", "الجنس", "تاريخ التسجيل"]
    rows = [[u.full_name_ar, u.email, u.get_role_display(), u.phone,
             {"male": "ذكر", "female": "أنثى"}.get(u.gender, "") if u.gender else "—", u.created_at.strftime("%Y-%m-%d")] for u in users]
    return _export_list_pdf(request, "التسجيلات الجديدة", headers, rows, "inscriptions_list")
@login_required
@role_required(User.Role.MAIN_ADMIN, User.Role.SUB_ADMIN)
def admin_inscriptions_export_excel(request):
    users = User.objects.filter(is_approved=User.ApprovalStatus.PENDING)
    headers = ["الاسم", "البريد", "الدور", "الهاتف", "الجنس", "تاريخ التسجيل"]
    rows = [[u.full_name_ar, u.email, u.get_role_display(), u.phone,
             {"male": "ذكر", "female": "أنثى"}.get(u.gender, "") if u.gender else "—", u.created_at.strftime("%Y-%m-%d")] for u in users]
    return _export_list_excel(request, "التسجيلات الجديدة", headers, rows, "inscriptions_list")
@login_required
@role_required(User.Role.MAIN_ADMIN, User.Role.SUB_ADMIN)
def admin_teacher_detail(request, pk):

    teacher = get_object_or_404(User, pk=pk, role=User.Role.TEACHER)
    _check_batch_access(request.user, teacher.batch_id)

    circles = Circle.objects.filter(teacher=teacher).annotate(
        active_students=Count('enrollments', filter=Q(enrollments__status=CircleEnrollment.Status.ACTIVE))
    )

    total_sessions = Session.objects.filter(circle__teacher=teacher).count()
    total_students = CircleEnrollment.objects.filter(
        circle__teacher=teacher, status=CircleEnrollment.Status.ACTIVE
    ).count()
    total_attendance = Attendance.objects.filter(session__circle__teacher=teacher).count()
    present_count = Attendance.objects.filter(
        session__circle__teacher=teacher, status=Attendance.Status.PRESENT
    ).count()
    attendance_rate = round((present_count / total_attendance * 100)) if total_attendance else 0

    recent_sessions = Session.objects.filter(circle__teacher=teacher).select_related(
        "circle"
    ).order_by("-session_date")[:5]

    recent_notifications = Notification.objects.filter(recipient=teacher).order_by("-created_at")[:5]

    return render(request, "dashboard/teachers/detail.html", {
        "teacher": teacher,
        "circles": circles,
        "total_sessions": total_sessions,
        "total_students": total_students,
        "attendance_rate": attendance_rate,
        "present_count": present_count,
        "recent_sessions": recent_sessions,
        "recent_notifications": recent_notifications,
    })
@login_required
@role_required(User.Role.MAIN_ADMIN, User.Role.SUB_ADMIN)
def admin_teacher_toggle_status(request, pk):

    teacher = get_object_or_404(User, pk=pk, role=User.Role.TEACHER)
    teacher.is_active = not teacher.is_active
    teacher.save(update_fields=["is_active"])
    return redirect("accounts:admin_teacher_detail", pk=pk)
@login_required
@role_required(User.Role.MAIN_ADMIN, User.Role.SUB_ADMIN)
def admin_teacher_create(request):

    if request.method == "POST":
        form = SignupForm(request.POST)
        if form.is_valid():
            user = form.save()
            user.role = User.Role.TEACHER
            user.is_approved = User.ApprovalStatus.APPROVED
            user.save()
            return redirect("accounts:admin_teachers")
    else:
        form = SignupForm()

    return render(request, "dashboard/teachers/create.html", {"form": form})
@login_required
@role_required(User.Role.MAIN_ADMIN, User.Role.SUB_ADMIN)
def admin_teacher_edit(request, pk):
    teacher = get_object_or_404(User, pk=pk, role=User.Role.TEACHER)

    # A Sub Admin may only edit teachers of a batch they supervise.
    _check_batch_access(request.user, teacher.batch_id)

    if request.method == "POST":
        teacher.full_name_ar = request.POST.get("full_name_ar", teacher.full_name_ar)
        teacher.email = request.POST.get("email", teacher.email)
        teacher.phone = request.POST.get("phone", teacher.phone)
        teacher.gender = request.POST.get("gender", teacher.gender)
        teacher.is_active = request.POST.get("is_active") == "true"
        if request.user.role == User.Role.MAIN_ADMIN:
            batch_id = request.POST.get("batch") or None
            if batch_id is None or Batch.objects.filter(pk=batch_id).exists():
                teacher.batch_id = batch_id
        teacher.save()

        circle_ids = request.POST.getlist("circles")
        Circle.objects.filter(teacher=teacher).update(teacher=None)
        if circle_ids:
            Circle.objects.filter(id__in=circle_ids, status=Circle.Status.ACTIVE).update(teacher=teacher)

        messages.success(request, "تم تحديث بيانات المعلم بنجاح")
        return redirect("accounts:admin_teachers")

    circles = scoping.scoped_circles(
        request.user, Circle.objects.filter(status=Circle.Status.ACTIVE)
    ).order_by("name")
    return render(request, "dashboard/teachers/edit.html", {
        "teacher": teacher,
        "circles": circles,
        "active_batches": Batch.objects.filter(status=Batch.Status.ACTIVE),
    })
@login_required
@role_required(User.Role.MAIN_ADMIN)
def admin_supervisors(request):

    search = request.GET.get("search", "")
    supervisors = User.objects.filter(role=User.Role.SUB_ADMIN).order_by('-created_at')

    if search:
        supervisors = supervisors.filter(Q(full_name_ar__icontains=search) | Q(email__icontains=search))

    paginator = Paginator(supervisors, 15)
    page_obj = paginator.get_page(request.GET.get('page', 1))

    return render(request, 'dashboard/supervisors/list.html', {'supervisors': page_obj, 'page_obj': page_obj})
@login_required
@role_required(User.Role.MAIN_ADMIN)
def admin_supervisor_create(request):

    if request.method == "POST":
        form = SignupForm(request.POST)
        if form.is_valid():
            user = form.save()
            user.role = User.Role.SUB_ADMIN
            user.is_approved = User.ApprovalStatus.APPROVED
            user.save()
            return redirect("accounts:admin_supervisors")
    else:
        form = SignupForm()

    return render(request, "dashboard/supervisors/create.html", {"form": form})


@login_required
@role_required(User.Role.MAIN_ADMIN)
def admin_supervisor_edit(request, pk):
    supervisor = get_object_or_404(User, pk=pk, role=User.Role.SUB_ADMIN)

    if request.method == "POST":
        form = ProfileForm(request.POST, instance=supervisor)
        if form.is_valid():
            supervisor = form.save(commit=False)
            supervisor.role = User.Role.SUB_ADMIN
            supervisor.is_active = request.POST.get("is_active") == "true"
            supervisor.save()
            messages.success(request, "تم تحديث بيانات المشرف بنجاح")
            return redirect("accounts:admin_supervisors")
    else:
        form = ProfileForm(instance=supervisor)

    return render(request, "dashboard/supervisors/edit.html", {
        "form": form,
        "supervisor": supervisor,
    })
@login_required
@role_required(User.Role.MAIN_ADMIN, User.Role.SUB_ADMIN)
def admin_circles(request):

    batch_ids = _scoped_batch_ids(request.user)
    search = request.GET.get("search", "")
    status_filter = request.GET.get("status", "")

    base_qs = Circle.objects.select_related('teacher')
    if batch_ids is not None:
        base_qs = base_qs.filter(batch_id__in=batch_ids)
    circles = (
        base_qs.annotate(
            active_students_count=Count('enrollments', filter=Q(enrollments__status=CircleEnrollment.Status.ACTIVE)),
            sessions_count=Count('sessions'),
        )
        .order_by('-created_at')
    )

    if status_filter:
        circles = circles.filter(status=status_filter)

    if search:
        circles = circles.filter(Q(name__icontains=search) | Q(teacher__full_name_ar__icontains=search) | Q(location__icontains=search))

    total_count = circles.count()
    active_count = Circle.objects.filter(status=Circle.Status.ACTIVE).count()
    paused_count = Circle.objects.filter(status=Circle.Status.PAUSED).count()
    inactive_count = Circle.objects.filter(status=Circle.Status.INACTIVE).count()

    paginator = Paginator(circles, 15)
    page_obj = paginator.get_page(request.GET.get('page', 1))

    return render(request, 'dashboard/circles/list.html', {
        'circles': page_obj,
        'page_obj': page_obj,
        'total_count': total_count,
        'active_count': active_count,
        'paused_count': paused_count,
        'inactive_count': inactive_count,
    })
@login_required
@role_required(User.Role.MAIN_ADMIN, User.Role.SUB_ADMIN)
def admin_circle_detail(request, pk):

    circle = get_object_or_404(
        Circle.objects.select_related('teacher'),
        pk=pk
    )
    _check_batch_access(request.user, circle.batch_id)

    if request.method == "POST":
        action = request.POST.get("action")
        if action == "add_student":
            student_id = request.POST.get("student_id")
            try:
                student = User.objects.get(pk=student_id, role=User.Role.STUDENT)
                CircleEnrollment._check_batch_match(student, circle)
                CircleEnrollment.objects.get_or_create(
                    circle=circle, student=student,
                    defaults={"status": CircleEnrollment.Status.ACTIVE},
                )
            except User.DoesNotExist:
                pass
            except ValidationError as e:
                messages.error(request, "؛ ".join(e.messages))
        elif action == "remove_student":
            enrollment_id = request.POST.get("enrollment_id")
            enrollment = get_object_or_404(
                CircleEnrollment, pk=enrollment_id, circle=circle,
                status=CircleEnrollment.Status.ACTIVE,
            )
            enrollment.status = CircleEnrollment.Status.INACTIVE
            enrollment.left_at = timezone.now()
            enrollment.save()
            Notification.objects.create(
                recipient=enrollment.student,
                type=Notification.Type.SYSTEM,
                title="تم إزالتك من الحلقة",
                message=f"لقد تم إزالتك من حلقة {circle.name}",
            )
            messages.success(request, f'تم إزالة {enrollment.student.full_name_ar} من الحلقة')
            return redirect("accounts:admin_circles")
        elif action == "approve_enrollment":
            enrollment_id = request.POST.get("enrollment_id")
            enrollment = get_object_or_404(
                CircleEnrollment, pk=enrollment_id, circle=circle,
                status=CircleEnrollment.Status.PENDING,
            )
            enrollment.status = CircleEnrollment.Status.ACTIVE
            enrollment.save()
            Notification.objects.create(
                recipient=enrollment.student,
                type=Notification.Type.SYSTEM,
                title="تم قبول طلب التسجيل",
                message=f"تم قبول طلب تسجيلك في حلقة {circle.name}",
                link=reverse("accounts:student_circle_detail", args=[circle.pk]),
            )
        elif action == "reject_enrollment":
            enrollment_id = request.POST.get("enrollment_id")
            enrollment = get_object_or_404(
                CircleEnrollment, pk=enrollment_id, circle=circle,
                status=CircleEnrollment.Status.PENDING,
            )
            enrollment.status = CircleEnrollment.Status.INACTIVE
            enrollment.left_at = timezone.now()
            enrollment.save()
            Notification.objects.create(
                recipient=enrollment.student,
                type=Notification.Type.SYSTEM,
                title="تم رفض طلب التسجيل",
                message=f"عذراً، لم يتم قبول طلب تسجيلك في حلقة {circle.name}",
            )
        elif action == "toggle_status":
            circle.status = Circle.Status.PAUSED if circle.status == Circle.Status.ACTIVE else Circle.Status.ACTIVE
            circle.save()
            status_display = circle.get_status_display()
            messages.success(request, f'تم تغيير حالة الحلقة إلى {status_display}')
            return redirect("accounts:admin_circle_detail", pk=circle.pk)
        elif action == "transfer_student":
            student_id = request.POST.get("student_id", "").strip()
            target_circle_id = request.POST.get("target_circle_id")
            try:
                student = User.objects.get(pk=student_id, role=User.Role.STUDENT)
                if not target_circle_id:
                    raise ValueError("target_circle_id is required")
                target_circle = get_object_or_404(Circle, pk=target_circle_id)
                source_enrollment = CircleEnrollment.objects.filter(
                    circle=circle, student=student, status=CircleEnrollment.Status.ACTIVE
                ).first()
                if not source_enrollment:
                    messages.error(request, 'الطالب غير مسجل في هذه الحلقة')
                elif target_circle == circle:
                    messages.error(request, 'لا يمكن نقل الطالب إلى نفس الحلقة')
                else:
                    CircleEnrollment._check_batch_match(student, target_circle)
                    source_enrollment.status = CircleEnrollment.Status.INACTIVE
                    source_enrollment.left_at = timezone.now()
                    source_enrollment.save()
                    enrollment, created = CircleEnrollment.objects.get_or_create(
                        circle=target_circle, student=student,
                        defaults={"status": CircleEnrollment.Status.ACTIVE},
                    )
                    if not created:
                        enrollment.status = CircleEnrollment.Status.ACTIVE
                        enrollment.left_at = None
                        enrollment.save()
                    messages.success(request, f'تم نقل {student.full_name_ar} إلى {target_circle.name}')
            except User.DoesNotExist:
                messages.error(request, 'الطالب غير موجود')
            except ValidationError as e:
                messages.error(request, "؛ ".join(e.messages))
            except (ValueError, Circle.DoesNotExist):
                messages.error(request, 'الحلقة الهدف غير موجودة')
            except Http404:
                messages.error(request, 'الحلقة الهدف غير موجودة')
            return redirect("accounts:admin_circle_detail", pk=circle.pk)
        elif action == "reject_enrollment_with_reason":
            enrollment_id = request.POST.get("enrollment_id")
            reason = request.POST.get("reason", "")
            enrollment = get_object_or_404(
                CircleEnrollment, pk=enrollment_id, circle=circle,
                status=CircleEnrollment.Status.PENDING,
            )
            enrollment.status = CircleEnrollment.Status.INACTIVE
            enrollment.left_at = timezone.now()
            enrollment.save()
            msg = f"عذراً، لم يتم قبول طلب تسجيلك في حلقة {circle.name}"
            if reason:
                msg += f". السبب: {reason}"
            Notification.objects.create(
                recipient=enrollment.student,
                type=Notification.Type.SYSTEM,
                title="تم رفض طلب التسجيل",
                message=msg,
            )
        return redirect("accounts:admin_circle_detail", pk=pk)

    pending_enrollments = circle.enrollments.filter(
        status=CircleEnrollment.Status.PENDING
    ).select_related('student').order_by('enrolled_at')

    active_enrollments = circle.enrollments.filter(
        status=CircleEnrollment.Status.ACTIVE
    ).select_related('student').order_by('-enrolled_at')

    past_enrollments = circle.enrollments.filter(
        status__in=[CircleEnrollment.Status.INACTIVE, CircleEnrollment.Status.DROPPED]
    ).select_related('student').order_by('-left_at')[:10]

    other_circles = Circle.objects.filter(status=Circle.Status.ACTIVE).exclude(pk=circle.pk)

    students_list = User.objects.filter(
        role=User.Role.STUDENT, is_approved=User.ApprovalStatus.APPROVED, is_active=True
    ).exclude(
        enrollments__in=circle.enrollments.filter(
            status__in=[CircleEnrollment.Status.ACTIVE, CircleEnrollment.Status.PENDING]
        )
    ).order_by('full_name_ar')[:50]

    sessions = Session.objects.filter(circle=circle).order_by('-session_date')[:10]
    sessions_count = Session.objects.filter(circle=circle).count()

    attendance_history = Session.objects.filter(circle=circle).prefetch_related(
        "attendance_records__student"
    ).order_by("-session_date")[:20]

    attendance_summary = Attendance.objects.filter(session__circle=circle).values(
        "session__session_date", "status"
    ).annotate(count=Count("id")).order_by("session__session_date")

    attendance_records = Attendance.objects.filter(session__circle=circle)
    total_attendance = attendance_records.count()
    present_count = attendance_records.filter(status=Attendance.Status.PRESENT).count()
    late_count = attendance_records.filter(status=Attendance.Status.LATE).count()
    absent_count = attendance_records.filter(status=Attendance.Status.ABSENT).count()
    attendance_rate = round((present_count + late_count) / total_attendance * 100) if total_attendance else 0

    hifz_total = MemorizationProgress.objects.filter(
        enrollment__circle=circle, type=MemorizationProgress.Type.HIFZ
    ).count()
    murajaa_total = MemorizationProgress.objects.filter(
        enrollment__circle=circle, type=MemorizationProgress.Type.MURAJAA
    ).count()

    return render(request, "dashboard/circles/detail.html", {
        "circle": circle,
        "active_enrollments": active_enrollments,
        "pending_enrollments": pending_enrollments,
        "pending_count": pending_enrollments.count(),
        "active_count": active_enrollments.count(),
        "past_enrollments": past_enrollments,
        "other_circles": other_circles,
        "students_list": students_list,
        "sessions": sessions,
        "sessions_count": sessions_count,
        "total_attendance": total_attendance,
        "present_count": present_count,
        "late_count": late_count,
        "absent_count": absent_count,
        "attendance_rate": attendance_rate,
        "hifz_total": hifz_total,
        "murajaa_total": murajaa_total,
        "attendance_history": attendance_history,
    })
@login_required
@role_required(User.Role.MAIN_ADMIN, User.Role.SUB_ADMIN)
def admin_circle_create(request):

    teachers = User.objects.filter(role=User.Role.TEACHER, is_approved=User.ApprovalStatus.APPROVED)
    own_batch = _sub_admin_batch(request.user)
    active_batches = Batch.objects.filter(status=Batch.Status.ACTIVE)

    if request.method == "POST":
        name = request.POST.get("name", "").strip()
        teacher_id = request.POST.get("teacher")
        location = request.POST.get("location", "")
        gender = request.POST.get("gender", Circle.Gender.MALE)
        max_students = request.POST.get("max_students", 30)
        schedule = request.POST.get("schedule", "")
        status = request.POST.get("status", Circle.Status.ACTIVE)
        circle_type = request.POST.get("circle_type", Circle.CircleType.HIFD)
        description = request.POST.get("description", "")
        batch_id = request.POST.get("batch") or None

        errors = {}
        if not name:
            errors["name"] = "اسم الحلقة مطلوب"
        if teacher_id:
            try:
                User.objects.get(pk=teacher_id, role=User.Role.TEACHER)
            except User.DoesNotExist:
                errors["teacher"] = "المعلم غير موجود"
        try:
            max_students = int(max_students)
            if max_students < 1:
                errors["max_students"] = "الحد الأقصى يجب أن يكون 1 على الأقل"
        except (ValueError, TypeError):
            errors["max_students"] = "قيمة غير صحيحة"
        if status not in dict(Circle.Status.choices):
            errors["status"] = "حالة غير صحيحة"
        if gender not in dict(Circle.Gender.choices):
            errors["gender"] = "جنس غير صحيح"
        if circle_type not in dict(Circle.CircleType.choices):
            errors["circle_type"] = "نوع حلقة غير صحيح"
        # الدفعة: a sub-admin's circle must land in a batch they supervise
        # (they may pick among several); the main admin picks any active
        # batch (or none). A sub-admin with no batches cannot create circles.
        supervised_ids = _scoped_batch_ids(request.user)
        if supervised_ids is not None:
            if not supervised_ids:
                errors["batch"] = "لا توجد دفعة تحت إشرافك"
            else:
                try:
                    chosen = int(batch_id) if batch_id else None
                except (TypeError, ValueError):
                    chosen = None
                if chosen not in supervised_ids:
                    batch_id = own_batch.pk if own_batch else None
        elif batch_id and not active_batches.filter(pk=batch_id).exists():
            errors["batch"] = "الدفعة غير موجودة"

        if errors:
            return render(request, "dashboard/circles/create.html", {
                "teachers": teachers,
                "active_batches": active_batches,
                "own_batch": own_batch,
                "errors": errors,
                "form_data": request.POST,
            })

        Circle.objects.create(
            teacher_id=teacher_id or None,
            name=name,
            location=location,
            gender=gender,
            circle_type=circle_type,
            max_students=max_students,
            schedule=schedule,
            status=status,
            description=description,
            batch_id=batch_id,
        )
        return redirect("accounts:admin_circles")

    return render(request, "dashboard/circles/create.html", {
        "teachers": teachers,
        "active_batches": active_batches,
        "own_batch": own_batch,
    })
@login_required
@role_required(User.Role.MAIN_ADMIN, User.Role.SUB_ADMIN)
def admin_requests(request):

    search = request.GET.get("search", "")
    req_type = request.GET.get("type", "")
    req_status = request.GET.get("status", "")
    req_priority = request.GET.get("priority", "")

    reqs = SupportRequest.objects.select_related('submitted_by').order_by('-created_at')

    if search:
        reqs = reqs.filter(Q(title__icontains=search) | Q(body__icontains=search) | Q(submitted_by__full_name_ar__icontains=search))
    if req_type:
        reqs = reqs.filter(type=req_type)
    if req_status:
        reqs = reqs.filter(status=req_status)
    if req_priority:
        reqs = reqs.filter(priority=req_priority)

    total_count = SupportRequest.objects.count()
    under_review_count = SupportRequest.objects.filter(status='under_review').count()
    approved_count = SupportRequest.objects.filter(status__in=['approved', 'resolved']).count()
    urgent_count = SupportRequest.objects.filter(priority__in=['urgent', 'high']).count()

    paginator = Paginator(reqs, 15)
    page_obj = paginator.get_page(request.GET.get('page', 1))

    return render(request, 'dashboard/requests/list.html', {
        'requests': page_obj,
        'page_obj': page_obj,
        'total_count': total_count,
        'under_review_count': under_review_count,
        'approved_count': approved_count,
        'urgent_count': urgent_count,
    })
@login_required
@role_required(User.Role.MAIN_ADMIN, User.Role.SUB_ADMIN)
def admin_request_detail(request, pk):

    request_obj = get_object_or_404(
        SupportRequest.objects.select_related('submitted_by'),
        pk=pk
    )

    if request.method == "POST":
        if request.POST.get("comment_body"):
            from apps.requests.models import Comment
            Comment.objects.create(
                request=request_obj,
                author=request.user,
                body=request.POST["comment_body"],
            )
            return redirect("accounts:admin_request_detail", pk=pk)

        new_status = request.POST.get("status")
        if new_status:
            from django.core.exceptions import ValidationError
            try:
                request_obj.transition_to(new_status, by=request.user)
            except ValidationError as e:
                messages.error(request, " ".join(e.messages))
        return redirect("accounts:admin_request_detail", pk=pk)

    comments = request_obj.comments.select_related("author").all()

    return render(request, "dashboard/requests/detail.html", {
        "request_obj": request_obj,
        "comments": comments,
    })
@login_required
@role_required(User.Role.MAIN_ADMIN, User.Role.SUB_ADMIN)
def admin_announcements(request):

    search = request.GET.get("search", "")
    announcements = Announcement.objects.select_related('author').order_by('-created_at')

    if search:
        announcements = announcements.filter(Q(title__icontains=search) | Q(body__icontains=search))

    paginator = Paginator(announcements, 15)
    page_obj = paginator.get_page(request.GET.get('page', 1))
    return render(request, 'dashboard/announcements/list.html', {'announcements': page_obj, 'page_obj': page_obj})
@login_required
@role_required(User.Role.MAIN_ADMIN, User.Role.SUB_ADMIN)
def admin_announcement_create(request):
    if request.method == "POST":
        title = request.POST.get("title", "").strip()
        body = request.POST.get("body", "").strip()
        if title and body:
            Announcement.objects.create(
                author=request.user,
                title=title,
                body=body,
            )
            return redirect("accounts:admin_announcements")
        return render(request, "dashboard/announcements/create.html", {"error": "يرجى ملء جميع الحقول"})

    return render(request, "dashboard/announcements/create.html")
@login_required
@role_required(User.Role.MAIN_ADMIN, User.Role.SUB_ADMIN)
def admin_notifications(request):

    from apps.notifications.models import Notification

    qs = Notification.objects.select_related("recipient").order_by("-created_at")

    search = request.GET.get("search", "")
    if search:
        qs = qs.filter(Q(title__icontains=search) | Q(message__icontains=search))

    is_read = request.GET.get("is_read", "")
    if is_read in ("true", "false"):
        qs = qs.filter(is_read=(is_read == "true"))

    notif_type = request.GET.get("type", "")
    if notif_type:
        qs = qs.filter(type=notif_type)

    paginator = Paginator(qs, 20)
    page_obj = paginator.get_page(request.GET.get("page", 1))

    return render(request, "dashboard/notifications/list.html", {
        "notifications": page_obj,
        "page_obj": page_obj,
    })
@login_required
@role_required(User.Role.MAIN_ADMIN, User.Role.SUB_ADMIN)
def admin_notification_create(request):

    from apps.notifications.models import Notification

    if request.method == "POST":
        notif_type = request.POST.get("type", "").strip()
        title = request.POST.get("title", "").strip()
        message = request.POST.get("message", "").strip()
        link = request.POST.get("link", "").strip()
        target = request.POST.get("target", "all").strip()

        if not (notif_type and title and message):
            return render(request, "dashboard/notifications/create.html", {
                "error": "يرجى ملء جميع الحقول المطلوبة",
                "form_data": request.POST,
            })

        if notif_type not in dict(Notification.Type.choices):
            return render(request, "dashboard/notifications/create.html", {
                "error": "نوع الإشعار غير صحيح",
                "form_data": request.POST,
            })

        role_map = {
            "admins": User.Role.MAIN_ADMIN,
            "supervisors": User.Role.SUB_ADMIN,
            "teachers": User.Role.TEACHER,
            "students": User.Role.STUDENT,
        }
        if target == "all":
            users = User.objects.filter(
                is_approved=User.ApprovalStatus.APPROVED, is_active=True
            )
        elif target in role_map:
            users = User.objects.filter(
                role=role_map[target],
                is_approved=User.ApprovalStatus.APPROVED,
                is_active=True,
            )
        else:
            return render(request, "dashboard/notifications/create.html", {
                "error": "الهدف غير صحيح",
                "form_data": request.POST,
            })

        notifications = [
            Notification(recipient=u, type=notif_type, title=title, message=message, link=link)
            for u in users
        ]
        Notification.objects.bulk_create(notifications, batch_size=500)
        return redirect("accounts:admin_notifications")

    return render(request, "dashboard/notifications/create.html")
@login_required
@role_required(User.Role.MAIN_ADMIN, User.Role.SUB_ADMIN)
def admin_reports(request):

    attendance_data = _report_attendance("")
    grades_data = _report_grades("")
    hifz_data = _report_hifz("")
    murajaa_data = _report_murajaa("")
    circles_data = _report_circles()

    circles = Circle.objects.filter(status=Circle.Status.ACTIVE)

    return render(request, 'dashboard/reports/index.html', {
        "attendance_summary": attendance_data,
        "grades_summary": grades_data,
        "hifz_summary": hifz_data,
        "murajaa_summary": murajaa_data,
        "circles_summary": circles_data,
        "circles": circles,
    })
@login_required
@role_required(User.Role.MAIN_ADMIN, User.Role.SUB_ADMIN)
def admin_report_data(request):

    report_type = request.GET.get("type", "empty")
    circle_id = request.GET.get("circle", "")

    if report_type == "attendance":
        data = _report_attendance(circle_id)
        return render(request, "dashboard/reports/partials/report_attendance.html", {"data": data})

    elif report_type == "grades":
        data = _report_grades(circle_id)
        return render(request, "dashboard/reports/partials/report_grades.html", {"data": data})

    elif report_type == "hifz":
        data = _report_hifz(circle_id)
        return render(request, "dashboard/reports/partials/report_hifz.html", {"data": data})

    elif report_type == "murajaa":
        data = _report_murajaa(circle_id)
        return render(request, "dashboard/reports/partials/report_murajaa.html", {"data": data})

    elif report_type == "circles":
        data = _report_circles()
        return render(request, "dashboard/reports/partials/report_circles.html", {"data": data})

    elif report_type == "teachers":
        data = _report_teachers()
        return render(request, "dashboard/reports/partials/report_teachers.html", {"data": data})

    return render(request, "dashboard/reports/partials/report_empty.html")


def _report_attendance(circle_id):
    total = Attendance.objects.count()
    present = Attendance.objects.filter(status__in=["present", "late", "left_early"]).count()
    absent = Attendance.objects.filter(status="absent").count()
    late = Attendance.objects.filter(status="late").count()
    rate = round(present / total * 100, 1) if total else 0

    circles = Circle.objects.filter(status="active").annotate(
        total_att=Count("sessions__attendance_records"),
        present_att=Count("sessions__attendance_records", filter=Q(
            sessions__attendance_records__status__in=["present", "late", "left_early"]
        )),
    )
    by_circle = []
    for c in circles:
        r = round(c.present_att / c.total_att * 100) if c.total_att else 0
        by_circle.append({
            "name": c.name,
            "present": c.present_att,
            "total": c.total_att,
            "rate": r,
        })

    daily_trend = []
    for i in range(13, -1, -1):
        d = date.today() - timedelta(days=i)
        day_total = Attendance.objects.filter(session__session_date=d).count()
        day_present = Attendance.objects.filter(
            session__session_date=d, status__in=["present", "late"]
        ).count()
        pct = round(day_present / day_total * 100) if day_total else 0
        daily_trend.append({
            "day": d.strftime("%d/%m"),
            "percentage": pct,
        })

    return {
        "present_count": present,
        "absent_count": absent,
        "late_count": late,
        "rate": rate,
        "by_circle": by_circle,
        "daily_trend": daily_trend,
    }


def _report_grades(circle_id):
    all_grades = RecitationGrade.objects.all()
    if circle_id:
        all_grades = all_grades.filter(session__circle_id=circle_id)

    percentages = all_grades.annotate(
        pct=ExpressionWrapper(
            F("score") * 100.0 / F("max_score"),
            output_field=FloatField()
        )
    )
    avg_all = percentages.aggregate(avg=Avg("pct"))["avg"] or 0
    excellent_count = percentages.filter(pct__gte=90).count()
    weak_count = percentages.filter(pct__lt=60).count()

    by_criterion = []
    for crit in EvaluationCriterion.objects.filter(is_active=True):
        crit_grades = percentages.filter(criterion=crit)
        avg = crit_grades.aggregate(avg=Avg("pct"))["avg"] or 0
        by_criterion.append({
            "name": crit.name_ar,
            "avg": round(avg, 1),
            "evaluations": crit_grades.count(),
        })

    top_students = (
        all_grades.values("student__full_name_ar")
        .annotate(avg=Avg(F("score") * 100.0 / F("max_score")))
        .order_by("-avg")[:5]
    )
    weak_students = (
        all_grades.values("student__full_name_ar")
        .annotate(avg=Avg(F("score") * 100.0 / F("max_score")))
        .order_by("avg")[:5]
    )

    return {
        "avg_all": round(avg_all, 1),
        "excellent_count": excellent_count,
        "weak_count": weak_count,
        "by_criterion": by_criterion,
        "top_students": [{"name": s["student__full_name_ar"], "avg": round(s["avg"], 1)} for s in top_students],
        "weak_students": [{"name": s["student__full_name_ar"], "avg": round(s["avg"], 1)} for s in weak_students],
    }


def _report_hifz(circle_id):
    qs = MemorizationProgress.objects.filter(type="hifz")
    if circle_id:
        qs = qs.filter(enrollment__circle_id=circle_id)

    _keys = thumn_start_keys()

    def _qs_thumns(sub_qs):
        return count_thumns(
            sub_qs.values_list("surah_id", "ayah_from", "ayah_to"), _keys=_keys
        )

    total_thumns = _qs_thumns(qs)
    mastered_thumns = _qs_thumns(qs.filter(status="mastered"))
    in_progress_thumns = _qs_thumns(qs.filter(status__in=["memorizing", "reviewed"]))
    mastery_rate = round(mastered_thumns / total_thumns * 100) if total_thumns else 0

    enrollments = CircleEnrollment.objects.filter(status="active")
    if circle_id:
        enrollments = enrollments.filter(circle_id=circle_id)

    students = []
    for en in enrollments.select_related("student", "circle", "current_surah"):
        hifz_records = MemorizationProgress.objects.filter(
            enrollment=en, type="hifz"
        )
        total = _qs_thumns(hifz_records)
        mastered = _qs_thumns(hifz_records.filter(status="mastered"))
        in_prog = total - mastered
        # Whole-mushaf completion: 480 athman = the full Quran.
        pct = round(mastered / TOTAL_THUMNS * 100) if total else 0

        students.append({
            "name": en.student.full_name_ar,
            "circle": en.circle.name,
            "total_thumns": total,
            "total_units": format_hizb_thumn(total),
            "mastered_thumns": mastered,
            "mastered_units": format_hizb_thumn(mastered),
            "in_progress_thumns": in_prog,
            "current_surah": en.current_surah.name_ar if en.current_surah else "—",
            "progress_pct": min(pct, 100),
        })

    students.sort(key=lambda x: x["mastered_thumns"], reverse=True)

    by_surah = []
    for surah in Surah.objects.all().order_by("id"):
        surah_hifz = qs.filter(surah=surah)
        if not surah_hifz.exists():
            continue
        memorized = surah_hifz.aggregate(
            t=Sum(F("ayah_to") - F("ayah_from") + 1)
        )["t"] or 0
        pct = round(memorized / surah.ayah_count * 100)
        by_surah.append({
            "id": surah.id,
            "name": surah.name_ar,
            "total": surah.ayah_count,
            "memorized": memorized,
            "percentage": min(pct, 100),
        })

    circles = Circle.objects.filter(status="active")

    return {
        "total_thumns": total_thumns,
        "total_units": format_hizb_thumn(total_thumns),
        "mastered_thumns": mastered_thumns,
        "mastered_units": format_hizb_thumn(mastered_thumns),
        "in_progress_thumns": in_progress_thumns,
        "in_progress_units": format_hizb_thumn(in_progress_thumns),
        "mastery_rate": mastery_rate,
        "students": students[:20],
        "by_surah": by_surah[:10],
        "circles": circles,
    }


def _report_murajaa(circle_id):
    qs = MemorizationProgress.objects.filter(type="murajaa")
    if circle_id:
        qs = qs.filter(enrollment__circle_id=circle_id)

    total_revisions = qs.count()
    mastered_count = qs.filter(status="mastered").count()
    weak_count = qs.filter(status="weak").count()

    status_labels = {
        "mastered": "متقن",
        "tested": "مختبر",
        "reviewed": "مراجع",
        "memorizing": "قيد المراجعة",
        "weak": "ضعيف",
    }
    status_distribution = []
    for key, label in status_labels.items():
        count = qs.filter(status=key).count()
        pct = round(count / total_revisions * 100) if total_revisions else 0
        status_distribution.append({"key": key, "label": label, "count": count, "percentage": pct})

    enrollments = CircleEnrollment.objects.filter(status="active")
    if circle_id:
        enrollments = enrollments.filter(circle_id=circle_id)

    total_students = enrollments.count()
    avg_revisions = round(total_revisions / total_students, 1) if total_students else 0

    students = []
    for en in enrollments.select_related("student", "circle"):
        murajaa_records = MemorizationProgress.objects.filter(
            enrollment=en, type="murajaa"
        )
        if not murajaa_records.exists():
            continue

        total_thumns = count_thumns(
            murajaa_records.values_list("surah_id", "ayah_from", "ayah_to")
        )
        mastered_thumns = count_thumns(
            murajaa_records.filter(status="mastered")
            .values_list("surah_id", "ayah_from", "ayah_to")
        )
        weak_thumns = count_thumns(
            murajaa_records.filter(status="weak")
            .values_list("surah_id", "ayah_from", "ayah_to")
        )
        revision_count = murajaa_records.aggregate(
            total=Sum("revision_count")
        )["total"] or 0
        last_revised = murajaa_records.order_by("-last_revised_at").first()
        last_rev_str = last_revised.last_revised_at.strftime("%d/%m/%Y") if last_revised and last_revised.last_revised_at else "—"

        mastered_pct = mastered_thumns / total_thumns * 100 if total_thumns else 0
        if mastered_pct >= 80:
            overall = "excellent"
        elif mastered_pct >= 60:
            overall = "good"
        elif mastered_pct >= 40:
            overall = "fair"
        else:
            overall = "weak"

        students.append({
            "name": en.student.full_name_ar,
            "circle": en.circle.name,
            "total_thumns": total_thumns,
            "total_units": format_hizb_thumn(total_thumns),
            "revision_count": revision_count,
            "mastered_thumns": mastered_thumns,
            "mastered_units": format_hizb_thumn(mastered_thumns),
            "weak_thumns": weak_thumns,
            "weak_units": format_hizb_thumn(weak_thumns),
            "last_revised": last_rev_str,
            "overall_status": overall,
        })

    students.sort(key=lambda x: x["mastered_thumns"], reverse=True)

    weak_students = sorted(students, key=lambda x: x["weak_thumns"], reverse=True)[:5]

    max_rev = max((s["revision_count"] for s in students), default=1) or 1
    frequency_chart = []
    for s in students[:15]:
        frequency_chart.append({
            "name": s["name"],
            "short_name": s["name"].split()[0][:8],
            "count": s["revision_count"],
            "height": round(s["revision_count"] / max_rev * 100),
        })

    mastery_rate = round(mastered_count / total_revisions * 100) if total_revisions else 0
    circles = Circle.objects.filter(status="active")

    return {
        "total_revisions": total_revisions,
        "mastered_count": mastered_count,
        "weak_count": weak_count,
        "avg_revisions": avg_revisions,
        "mastery_rate": mastery_rate,
        "status_distribution": status_distribution,
        "students": students[:20],
        "weak_students": weak_students,
        "frequency_chart": frequency_chart,
        "circles": circles,
    }
@login_required
@role_required(User.Role.MAIN_ADMIN, User.Role.SUB_ADMIN)
def admin_report_export_pdf(request):

    from io import BytesIO

    report_type = request.GET.get("type", "attendance")
    circle_id = request.GET.get("circle", "")

    report_titles = {
        "attendance": "تقرير الحضور",
        "grades": "تقرير الدرجات",
        "hifz": "تقرير الحفظ",
        "murajaa": "تقرير المراجعة",
        "circles": "تقرير الحلقات",
        "teachers": "تقرير المعلمين",
    }
    title = report_titles.get(report_type, "تقرير")

    pdf = _make_branded_pdf("P", title,
                            f"تاريخ التصدير: {date.today().strftime('%Y-%m-%d')}")
    pdf.add_page()

    if report_type == "attendance":
        data = _report_attendance(circle_id)
        pdf.set_font("Arabic", "", 12)
        pdf.cell(0, 8, f"إجمالي الحضور: {data['present_count']}    الغياب: {data['absent_count']}    التأخير: {data['late_count']}", align="C")
        pdf.ln(10)
        pdf.set_font("Arabic", "B", 14)
        pdf.cell(0, 8, f"نسبة الحضور: {data['rate']}%", align="C")
        pdf.ln(14)
        pdf.set_font("Arabic", "B", 12)
        pdf.cell(0, 8, "تفاصيل الحضور حسب الحلقة", align="C")
        pdf.ln(12)
        pdf.set_font("Arabic", "", 10)
        for c in data["by_circle"]:
            pdf.cell(0, 7, f"{c['name']}: {c['present']} من {c['total']} ({c['rate']}%)", align="C")
            pdf.ln(7)

    elif report_type == "grades":
        data = _report_grades(circle_id)
        pdf.set_font("Arabic", "", 12)
        pdf.cell(0, 8, f"المعدل العام: {data['avg_all']}%", align="C")
        pdf.ln(8)
        pdf.cell(0, 8, f"المتفوقون (90% فأكثر): {data['excellent_count']}    الضعاف (أقل من 60%): {data['weak_count']}", align="C")
        pdf.ln(14)
        pdf.set_font("Arabic", "B", 12)
        pdf.cell(0, 8, "الدرجات حسب معيار التقييم", align="C")
        pdf.ln(12)
        pdf.set_font("Arabic", "", 10)
        for c in data["by_criterion"]:
            pdf.cell(0, 7, f"{c['name']}: {c['avg']}%  ({c['evaluations']} تقييم)", align="C")
            pdf.ln(7)

    elif report_type == "hifz":
        data = _report_hifz(circle_id)
        pdf.set_font("Arabic", "", 12)
        pdf.cell(0, 8, f"إجمالي المحفوظ: {data['total_units']} ({data['total_thumns']} ثمن)", align="C")
        pdf.ln(8)
        pdf.cell(0, 8, f"المتقن: {data['mastered_units']}    قيد الحفظ: {data['in_progress_units']}", align="C")
        pdf.ln(10)
        pdf.set_font("Arabic", "B", 14)
        pdf.cell(0, 8, f"نسبة الإتقان: {data['mastery_rate']}%", align="C")

    elif report_type == "murajaa":
        data = _report_murajaa(circle_id)
        pdf.set_font("Arabic", "", 12)
        pdf.cell(0, 8, f"إجمالي المراجعات: {data['total_revisions']}", align="C")
        pdf.ln(8)
        pdf.cell(0, 8, f"مراجعات متقنة: {data['mastered_count']}    مراجعات ضعيفة: {data['weak_count']}", align="C")
        pdf.ln(8)
        pdf.cell(0, 8, f"متوسط المراجعات لكل طالب: {data['avg_revisions']}", align="C")
        pdf.ln(10)
        pdf.set_font("Arabic", "B", 14)
        pdf.cell(0, 8, f"نسبة إتقان المراجعة: {data['mastery_rate']}%", align="C")

    elif report_type == "circles":
        data = _report_circles()
        pdf.set_font("Arabic", "", 12)
        pdf.cell(0, 8, f"إجمالي الحلقات: {data['total']}    النشطة: {data['active']}", align="C")
        pdf.ln(8)
        pdf.cell(0, 8, f"متوسط عدد الطلاب لكل حلقة: {data['avg_students']}    إجمالي الحصص: {data['total_sessions']}", align="C")

    elif report_type == "teachers":
        data = _report_teachers()
        pdf.set_font("Arabic", "", 12)
        pdf.cell(0, 8, f"إجمالي المعلمين: {data['total_teachers']}", align="C")
        pdf.ln(8)
        pdf.cell(0, 8, f"متوسط نسبة الحضور: {data['avg_attendance']}%    متوسط الدرجات: {data['avg_grade']}%", align="C")

    buf = BytesIO()
    pdf.output(buf)
    buf.seek(0)
    response = HttpResponse(buf.read(), content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="report_{report_type}.pdf"'
    return response
@login_required
@role_required(User.Role.MAIN_ADMIN, User.Role.SUB_ADMIN)
def admin_report_export_excel(request):

    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill

    report_type = request.GET.get("type", "attendance")
    circle_id = request.GET.get("circle", "")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "تقرير"

    header_font = Font(name="Noto Sans Arabic", size=14, bold=True)
    cell_font = Font(name="Noto Sans Arabic", size=10)
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_font_white = Font(name="Noto Sans Arabic", size=10, bold=True, color="FFFFFF")

    def style_header(ws, row, cols):
        for col in range(1, cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
    ws.cell(row=1, column=1, value=f"تقرير {report_type}").font = header_font

    if report_type == "attendance":
        data = _report_attendance(circle_id)
        ws.cell(row=3, column=1, value=f"إجمالي الحضور: {data['present_count']}")
        ws.cell(row=4, column=1, value=f"الغياب: {data['absent_count']}")
        ws.cell(row=5, column=1, value=f"التأخير: {data['late_count']}")
        ws.cell(row=6, column=1, value=f"نسبة الحضور: {data['rate']}%")
        headers = ["الحلقة", "حاضر", "الإجمالي", "النسبة"]
        row = 8
        for i, h in enumerate(headers, 1):
            ws.cell(row=row, column=i, value=h)
        style_header(ws, row, len(headers))
        for c in data["by_circle"]:
            row += 1
            ws.cell(row=row, column=1, value=c["name"])
            ws.cell(row=row, column=2, value=c["present"])
            ws.cell(row=row, column=3, value=c["total"])
            ws.cell(row=row, column=4, value=f"{c['rate']}%")

    elif report_type == "grades":
        data = _report_grades(circle_id)
        ws.cell(row=3, column=1, value=f"المعدل العام: {data['avg_all']}%")
        ws.cell(row=4, column=1, value=f"المتفوقون: {data['excellent_count']}")
        ws.cell(row=5, column=1, value=f"الضعاف: {data['weak_count']}")
        headers = ["المعيار", "المتوسط", "عدد التقييمات"]
        row = 7
        for i, h in enumerate(headers, 1):
            ws.cell(row=row, column=i, value=h)
        style_header(ws, row, len(headers))
        for c in data["by_criterion"]:
            row += 1
            ws.cell(row=row, column=1, value=c["name"])
            ws.cell(row=row, column=2, value=f"{c['avg']}%")
            ws.cell(row=row, column=3, value=c["evaluations"])

    elif report_type == "hifz":
        data = _report_hifz(circle_id)
        ws.cell(row=3, column=1, value=f"إجمالي المحفوظ: {data['total_units']} ({data['total_thumns']} ثمن)")
        ws.cell(row=4, column=1, value=f"المُتقن: {data['mastered_units']}")
        ws.cell(row=5, column=1, value=f"قيد الحفظ: {data['in_progress_units']}")
        ws.cell(row=6, column=1, value=f"نسبة الإتقان: {data['mastery_rate']}%")

    elif report_type == "murajaa":
        data = _report_murajaa(circle_id)
        ws.cell(row=3, column=1, value=f"إجمالي المراجعات: {data['total_revisions']}")
        ws.cell(row=4, column=1, value=f"المُتقن: {data['mastered_count']}")
        ws.cell(row=5, column=1, value=f"الضعيف: {data['weak_count']}")
        ws.cell(row=6, column=1, value=f"نسبة الإتقان: {data['mastery_rate']}%")

    elif report_type == "circles":
        data = _report_circles()
        ws.cell(row=3, column=1, value=f"إجمالي الحلقات: {data['total']}")
        ws.cell(row=4, column=1, value=f"النشطة: {data['active']}")
        ws.cell(row=5, column=1, value=f"متوسط الطلاب: {data['avg_students']}")
        headers = ["الحلقة", "المعلم", "الطلاب", "الحصص", "نسبة الحضور", "المعدل", "المحفوظ"]
        row = 7
        for i, h in enumerate(headers, 1):
            ws.cell(row=row, column=i, value=h)
        style_header(ws, row, len(headers))
        for c in data["circles"]:
            row += 1
            ws.cell(row=row, column=1, value=c["name"])
            ws.cell(row=row, column=2, value=c["teacher"])
            ws.cell(row=row, column=3, value=c["students"])
            ws.cell(row=row, column=4, value=c["sessions"])
            ws.cell(row=row, column=5, value=f"{c['attendance_rate']}%")
            ws.cell(row=row, column=6, value=f"{c['avg_grade']}%")
            ws.cell(row=row, column=7, value=c["memorized_units"])

    elif report_type == "teachers":
        data = _report_teachers()
        ws.cell(row=3, column=1, value=f"إجمالي المعلمين: {data['total_teachers']}")
        ws.cell(row=4, column=1, value=f"متوسط الحضور: {data['avg_attendance']}%")
        ws.cell(row=5, column=1, value=f"متوسط الدرجات: {data['avg_grade']}%")
        headers = ["المعلم", "الحلقات", "الطلاب", "الحصص", "الحضور", "المعدل"]
        row = 7
        for i, h in enumerate(headers, 1):
            ws.cell(row=row, column=i, value=h)
        style_header(ws, row, len(headers))
        for t in data["teachers"]:
            row += 1
            ws.cell(row=row, column=1, value=t["name"])
            ws.cell(row=row, column=2, value=t["circles"])
            ws.cell(row=row, column=3, value=t["students"])
            ws.cell(row=row, column=4, value=t["sessions"])
            ws.cell(row=row, column=5, value=f"{t['attendance_rate']}%")
            ws.cell(row=row, column=6, value=f"{t['avg_grade']}%")

    for col_cells in ws.columns:
        max_len = 0
        for cell in col_cells:
            if isinstance(cell, openpyxl.cell.cell.MergedCell):
                continue
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        # col_cells[0] might be a MergedCell if the first cell in column is merged
        first_cell = col_cells[0]
        if not isinstance(first_cell, openpyxl.cell.cell.MergedCell):
            ws.column_dimensions[first_cell.column_letter].width = min(max_len + 4, 40)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="report_{report_type}.xlsx"'
    wb.save(response)
    return response


def _report_circles():
    circles = Circle.objects.select_related("teacher").annotate(
        student_count=Count("enrollments", filter=Q(enrollments__status="active")),
        session_count=Count("sessions"),
    ).order_by("-student_count")

    result_circles = []
    total_sessions = 0
    for c in circles:
        total_sessions += c.session_count

        att_total = Attendance.objects.filter(session__circle=c).count()
        att_present = Attendance.objects.filter(
            session__circle=c, status__in=["present", "late", "left_early"]
        ).count()
        att_rate = round(att_present / att_total * 100) if att_total else 0

        avg_grade = RecitationGrade.objects.filter(session__circle=c).annotate(
            pct=F("score") * 100.0 / F("max_score")
        ).aggregate(avg=Avg("pct"))["avg"] or 0

        mem_thumns = count_thumns(
            MemorizationProgress.objects.filter(
                enrollment__circle=c, type="hifz", status="mastered"
            ).values_list("surah_id", "ayah_from", "ayah_to")
        )

        result_circles.append({
            "name": c.name,
            "teacher": c.teacher.full_name_ar if c.teacher else "—",
            "students": c.student_count,
            "sessions": c.session_count,
            "attendance_rate": att_rate,
            "avg_grade": round(avg_grade, 1),
            "memorized_thumns": mem_thumns,
            "memorized_units": format_hizb_thumn(mem_thumns),
        })

    return {
        "total": circles.count(),
        "active": circles.filter(status="active").count(),
        "avg_students": round(
            sum(c["students"] for c in result_circles) / max(len(result_circles), 1), 1
        ),
        "total_sessions": total_sessions,
        "circles": result_circles,
    }


def _report_teachers():
    teachers = User.objects.filter(role=User.Role.TEACHER, is_approved=User.ApprovalStatus.APPROVED).annotate(
        circles_count=Count("teaching_circles", filter=Q(teaching_circles__status=Circle.Status.ACTIVE)),
        students_count=Count("teaching_circles__enrollments", filter=Q(
            teaching_circles__enrollments__status=CircleEnrollment.Status.ACTIVE
        )),
        sessions_count=Count("teaching_circles__sessions"),
    ).order_by("-students_count")

    result = []
    for t in teachers:
        attendance_qs = Attendance.objects.filter(session__circle__teacher=t)
        total_att = attendance_qs.count()
        present_att = attendance_qs.filter(status__in=["present", "late"]).count()
        att_rate = round(present_att / total_att * 100) if total_att else 0

        avg_grade = RecitationGrade.objects.filter(session__circle__teacher=t).annotate(
            pct=F("score") * 100.0 / F("max_score")
        ).aggregate(avg=Avg("pct"))["avg"] or 0

        result.append({
            "name": t.full_name_ar,
            "circles": t.circles_count,
            "students": t.students_count,
            "sessions": t.sessions_count,
            "attendance_rate": att_rate,
            "avg_grade": round(avg_grade, 1),
        })

    total_teachers = len(result)
    avg_attendance = round(
        sum(t["attendance_rate"] for t in result) / max(total_teachers, 1)
    )
    avg_grade_all = round(
        sum(t["avg_grade"] for t in result) / max(total_teachers, 1), 1
    )

    return {
        "total_teachers": total_teachers,
        "avg_attendance": avg_attendance,
        "avg_grade": avg_grade_all,
        "teachers": result,
    }
@login_required
@role_required(User.Role.MAIN_ADMIN, User.Role.SUB_ADMIN)
def report_exam_results(request):
    marks = ExamMark.objects.filter(
        status=ExamMark.Status.APPROVED,
        exam__status=Exam.Status.COMPLETED,
    ).select_related("exam__circle", "student", "approved_by").order_by("-exam__exam_date")
    students_data = {}
    for m in marks:
        sid = m.student_id
        if sid not in students_data:
            students_data[sid] = {
                "student": m.student,
                "total_exams": 0,
                "percentages": [],
                "last_exam_date": m.exam.exam_date,
                "last_circle": m.exam.circle.name if m.exam.circle else None,
            }
        sd = students_data[sid]
        sd["total_exams"] += 1
        sd["percentages"].append(m.percentage)
        if m.exam.exam_date >= sd["last_exam_date"]:
            sd["last_exam_date"] = m.exam.exam_date
            sd["last_circle"] = m.exam.circle.name if m.exam.circle else sd["last_circle"]
    exam_data = []
    for sid, sd in students_data.items():
        pcts = sd["percentages"]
        exam_data.append({
            "student": sd["student"],
            "total_exams": sd["total_exams"],
            "avg_score": round(sum(pcts) / len(pcts), 1),
            "last_exam_date": sd["last_exam_date"],
            "last_circle": sd["last_circle"],
        })
    exam_data.sort(key=lambda x: x["avg_score"], reverse=True)
    return render(request, "dashboard/reports/exam_results.html", {
        "exam_data": exam_data,
        "total_students": len(exam_data),
        "overall_avg": round(sum(e["avg_score"] for e in exam_data) / max(len(exam_data), 1), 1),
    })
@login_required
@role_required(User.Role.MAIN_ADMIN, User.Role.SUB_ADMIN)
def admin_exam_list(request):
    batch_ids = _scoped_batch_ids(request.user)
    exams = Exam.objects.select_related("circle", "created_by", "assigned_teacher").annotate(
        student_count_annotated=Count("marks"),
        average_marks_annotated=Avg("marks__marks_obtained"),
        approved_count=Count("marks", filter=Q(marks__status=ExamMark.Status.APPROVED)),
        pending_count=Count("marks", filter=Q(marks__status=ExamMark.Status.PENDING)),
        rejected_count=Count("marks", filter=Q(marks__status=ExamMark.Status.REJECTED)),
    )
    if batch_ids is not None:
        exams = exams.filter(circle__batch_id__in=batch_ids)
    return render(request, "dashboard/exams/list.html", {"exams": exams})
@login_required
@role_required(User.Role.MAIN_ADMIN, User.Role.SUB_ADMIN)
def admin_exam_create(request):
    circles = Circle.objects.filter(status=Circle.Status.ACTIVE)
    teachers = User.objects.filter(role=User.Role.TEACHER, is_approved=User.ApprovalStatus.APPROVED, is_active=True)
    if request.method == "POST":
        title = request.POST.get("title", "").strip()
        if not title:
            return render(request, "dashboard/exams/create.html", {
                "circles": circles, "teachers": teachers, "error": "عنوان الامتحان مطلوب"
            })
        data = {
            "title": title,
            "description": request.POST.get("description", ""),
            "exam_type": request.POST.get("exam_type", "monthly"),
            "circle_id": request.POST.get("circle") or None,
            "assigned_teacher_id": request.POST.get("assigned_teacher") or None,
            "exam_date": request.POST.get("exam_date", timezone.now().date()),
            "max_marks": float(request.POST.get("max_marks", 100)),
            "pass_percentage": float(request.POST.get("pass_percentage", 50)),
            "auto_publish": request.POST.get("auto_publish") == "on",
        }
        exam = create_exam(data, request.user)
        messages.success(request, "تم إنشاء الامتحان بنجاح")
        return redirect("accounts:admin_exam_detail", pk=exam.pk)
    return render(request, "dashboard/exams/create.html", {"circles": circles, "teachers": teachers})
@login_required
@role_required(User.Role.MAIN_ADMIN, User.Role.SUB_ADMIN)
def admin_exam_edit(request, pk):
    exam = get_object_or_404(Exam, pk=pk)
    circles = Circle.objects.filter(status=Circle.Status.ACTIVE)
    teachers = User.objects.filter(role=User.Role.TEACHER, is_approved=User.ApprovalStatus.APPROVED, is_active=True)
    if request.method == "POST":
        exam.title = request.POST.get("title", exam.title)
        exam.description = request.POST.get("description", "")
        exam.exam_type = request.POST.get("exam_type", exam.exam_type)
        exam.circle_id = request.POST.get("circle") or None
        exam.assigned_teacher_id = request.POST.get("assigned_teacher") or None
        exam.exam_date = request.POST.get("exam_date", exam.exam_date)
        exam.max_marks = float(request.POST.get("max_marks", 100))
        exam.pass_percentage = float(request.POST.get("pass_percentage", 50))
        exam.save()
        messages.success(request, "تم تحديث الامتحان")
        return redirect("accounts:admin_exam_detail", pk=exam.pk)
    return render(request, "dashboard/exams/edit.html", {
        "exam": exam, "circles": circles, "teachers": teachers,
    })
@login_required
@role_required(User.Role.MAIN_ADMIN, User.Role.SUB_ADMIN)
def admin_exam_detail(request, pk):
    exam = get_object_or_404(
        Exam.objects.select_related("circle", "created_by", "assigned_teacher"),
        pk=pk,
    )
    if exam.circle:
        _check_batch_access(request.user, exam.circle.batch_id)
    marks = exam.marks.select_related("student", "entered_by", "approved_by").all()
    enrolled_students = []
    if exam.circle:
        enrolled_students = list(User.objects.filter(
            enrollments__circle=exam.circle,
            enrollments__status=CircleEnrollment.Status.ACTIVE,
            role=User.Role.STUDENT,
        ).distinct())
    else:
        enrolled_students = list(User.objects.filter(
            role=User.Role.STUDENT, is_approved=User.ApprovalStatus.APPROVED
        ))
    existing_marks = {m.student_id: m for m in marks}
    history = exam.approval_history.select_related("performed_by").all()[:20]
    approval_progress = exam.approval_progress()
    return render(request, "dashboard/exams/detail.html", {
        "exam": exam, "marks": marks,
        "enrolled_students": enrolled_students,
        "existing_marks": existing_marks,
        "history": history,
        "approval_progress": approval_progress,
    })
@login_required
@role_required(User.Role.MAIN_ADMIN, User.Role.SUB_ADMIN)
def admin_exam_delete(request, pk):
    exam = get_object_or_404(Exam, pk=pk)
    exam.delete()
    messages.success(request, "تم حذف الامتحان")
    return redirect("accounts:admin_exam_list")
@login_required
@role_required(User.Role.MAIN_ADMIN, User.Role.SUB_ADMIN)
def admin_exam_publish(request, pk):
    exam = get_object_or_404(Exam, pk=pk)
    exam.status = Exam.Status.PUBLISHED
    exam.save(update_fields=["status"])
    notify_published(exam, request.user)
    messages.success(request, "تم نشر الامتحان")
    return redirect("accounts:admin_exam_detail", pk=exam.pk)
@login_required
@role_required(User.Role.MAIN_ADMIN, User.Role.SUB_ADMIN)
def admin_exam_approve_all(request, pk):
    exam = get_object_or_404(Exam, pk=pk)
    approve_all_marks(exam, request.user)
    messages.success(request, "تم اعتماد جميع النتائج")
    return redirect("accounts:admin_exam_detail", pk=exam.pk)
@login_required
@role_required(User.Role.MAIN_ADMIN, User.Role.SUB_ADMIN)
def admin_exam_reject_marks(request, pk):
    exam = get_object_or_404(Exam, pk=pk)
    if request.method != "POST":
        messages.error(request, "طلب غير صالح")
        return redirect("accounts:admin_exam_detail", pk=exam.pk)
    mark_ids = request.POST.getlist("mark_ids")
    reason = request.POST.get("reason", "")
    if not mark_ids:
        messages.error(request, "لم يتم تحديد درجات للرفض")
        return redirect("accounts:admin_exam_detail", pk=exam.pk)
    reject_marks(exam, [int(m) for m in mark_ids], request.user, reason)
    messages.success(request, "تم رفض الدرجات المحددة")
    return redirect("accounts:admin_exam_detail", pk=exam.pk)
@login_required
@role_required(User.Role.MAIN_ADMIN, User.Role.SUB_ADMIN)
def admin_exam_export_pdf(request, pk):
    from apps.exams.services import get_export_data
    from apps.exams.utils import generate_exam_pdf
    exam = get_object_or_404(Exam, pk=pk)
    export_data = get_export_data(exam)
    pdf_bytes = generate_exam_pdf(export_data)
    from django.http import HttpResponse
    response = HttpResponse(pdf_bytes, content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="exam_{exam.exam_code}.pdf"'
    return response
@login_required
@role_required(User.Role.MAIN_ADMIN, User.Role.SUB_ADMIN)
def admin_exam_export_csv(request, pk):
    from apps.exams.services import get_export_data
    from apps.exams.utils import generate_exam_csv
    exam = get_object_or_404(Exam, pk=pk)
    export_data = get_export_data(exam)
    csv_str = generate_exam_csv(export_data)
    from django.http import HttpResponse
    response = HttpResponse(csv_str, content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = f'attachment; filename="exam_{exam.exam_code}.csv"'
    return response


@login_required
@role_required(User.Role.MAIN_ADMIN, User.Role.SUB_ADMIN)
def admin_private_sessions(request):
    """Read-only oversight of every teacher↔student private (1-on-1) تسميع
    session across the platform, with an optional status filter."""
    from apps.memorization.models import PrivateSession
    sessions = PrivateSession.objects.select_related(
        "teacher", "student", "circle",
    ).order_by("-scheduled_date", "-created_at")
    status_filter = request.GET.get("status", "")
    if status_filter:
        sessions = sessions.filter(status=status_filter)
    paginator = Paginator(sessions, 25)
    page_obj = paginator.get_page(request.GET.get("page", 1))
    return render(request, "dashboard/admin/private_sessions.html", {
        "sessions": page_obj,
        "page_obj": page_obj,
        "status_filter": status_filter,
    })


# ── Batch Management ───────────────────────────────────────────────────


def _batch_detail_context(request, batch, batch_form=None):
    circles = Circle.objects.filter(batch=batch).select_related("teacher")
    unassigned_circles = Circle.objects.filter(batch__isnull=True).select_related("teacher").order_by("name")
    teachers = User.objects.filter(
        role=User.Role.TEACHER, batch=batch,
    ).order_by("full_name_ar")
    students_qs = User.objects.filter(
        role=User.Role.STUDENT, batch=batch,
    ).order_by("full_name_ar")

    students_paginator = Paginator(students_qs, 20)
    students_page = students_paginator.get_page(request.GET.get("page", 1))

    unassigned_users = User.objects.filter(
        is_approved=User.ApprovalStatus.APPROVED,
        role__in=[User.Role.STUDENT, User.Role.TEACHER],
    ).exclude(batch=batch).order_by("role", "full_name_ar")

    supervisors = (
        User.objects.filter(
            role=User.Role.SUB_ADMIN,
            is_approved=User.ApprovalStatus.APPROVED,
        )
        if request.user.role == User.Role.MAIN_ADMIN else User.objects.none()
    )

    return {
        "batch": batch,
        "batch_form": batch_form or BatchForm(instance=batch),
        "circles": circles,
        "unassigned_circles": unassigned_circles,
        "teachers": teachers,
        "students": students_page,
        "teachers_total": teachers.count(),
        "students_total": students_paginator.count,
        "page_obj": students_page,
        "unassigned_users": unassigned_users,
        "supervisors": supervisors,
        "status_choices": Batch.Status.choices,
    }


@login_required
@role_required(User.Role.MAIN_ADMIN, User.Role.SUB_ADMIN)
def admin_batch_list(request):
    batch_form = BatchForm()

    if request.method == "POST":
        if request.user.role != User.Role.MAIN_ADMIN:
            raise PermissionDenied
        batch_form = BatchForm(request.POST)
        if batch_form.is_valid():
            batch = batch_form.save(commit=False)
            batch.created_by = request.user
            batch.save()
            batch_form.save_m2m()
            messages.success(request, f"تم إنشاء الدفعة {batch.name} بنجاح")
            return redirect("accounts:admin_batch_detail", pk=batch.pk)

    if request.user.role == User.Role.MAIN_ADMIN:
        batches = Batch.objects.all()
    else:
        batches = Batch.objects.filter(Q(sub_admin=request.user) | Q(sub_admins=request.user))

    status_filter = request.GET.get("status", "")
    search = request.GET.get("search", "")
    if status_filter in dict(Batch.Status.choices):
        batches = batches.filter(status=status_filter)
    if search:
        batches = batches.filter(Q(name__icontains=search) | Q(year__icontains=search))

    batches = batches.select_related("sub_admin").annotate(
        students_count=Count("users", filter=Q(users__role=User.Role.STUDENT), distinct=True),
        teachers_count=Count("users", filter=Q(users__role=User.Role.TEACHER), distinct=True),
        circles_count=Count("circles", distinct=True),
    ).order_by("-created_at")

    return render(request, "dashboard/admin/batches/list.html", {
        "batches": batches,
        "batch_form": batch_form,
        "status_filter": status_filter,
        "search": search,
    })


@login_required
@role_required(User.Role.MAIN_ADMIN)
def admin_batch_create(request):
    if request.method == "POST":
        form = BatchForm(request.POST)
        if form.is_valid():
            batch = form.save(commit=False)
            batch.created_by = request.user
            batch.save()
            form.save_m2m()
            messages.success(request, f"تم إنشاء الدفعة {batch.name} بنجاح")
            return redirect("accounts:admin_batch_detail", pk=batch.pk)

        batches = Batch.objects.select_related("sub_admin").annotate(
            students_count=Count("users", filter=Q(users__role=User.Role.STUDENT), distinct=True),
            teachers_count=Count("users", filter=Q(users__role=User.Role.TEACHER), distinct=True),
            circles_count=Count("circles", distinct=True),
        ).order_by("-created_at")
        return render(request, "dashboard/admin/batches/list.html", {
            "batches": batches,
            "batch_form": form,
            "status_filter": "",
            "search": "",
        })

    return redirect("accounts:admin_batch_list")


@login_required
@role_required(User.Role.MAIN_ADMIN)
def admin_batch_edit(request, pk):
    batch = get_object_or_404(Batch, pk=pk)
    if request.method == "POST":
        form = BatchForm(request.POST, instance=batch)
        if form.is_valid():
            form.save()
            messages.success(request, f"تم تحديث الدفعة {batch.name} بنجاح")
            return redirect("accounts:admin_batch_detail", pk=batch.pk)
        return render(
            request,
            "dashboard/admin/batches/detail.html",
            _batch_detail_context(request, batch, form),
        )

    return redirect("accounts:admin_batch_detail", pk=batch.pk)


@login_required
@role_required(User.Role.MAIN_ADMIN, User.Role.SUB_ADMIN)
def admin_batch_detail(request, pk):
    """The one page for everything دفعة: members, circles, supervisor,
    status and deletion are all actions of this view — no satellite pages."""
    batch = get_object_or_404(
        Batch.objects.select_related("sub_admin").prefetch_related("sub_admins"),
        pk=pk,
    )
    if (
        request.user.role != User.Role.MAIN_ADMIN
        and batch.sub_admin != request.user
        and not batch.sub_admins.filter(pk=request.user.pk).exists()
    ):
        raise PermissionDenied

    if request.method == "POST":
        action = request.POST.get("action")
        is_main_admin = request.user.role == User.Role.MAIN_ADMIN

        if action == "update_batch" and is_main_admin:
            form = BatchForm(request.POST, instance=batch)
            if form.is_valid():
                form.save()
                messages.success(request, f"تم تحديث الدفعة {batch.name} بنجاح")
                return redirect("accounts:admin_batch_detail", pk=batch.pk)
            return render(
                request,
                "dashboard/admin/batches/detail.html",
                _batch_detail_context(request, batch, form),
            )

        elif action == "assign_users":
            user_ids = request.POST.getlist("user_ids")
            candidates = User.objects.filter(
                pk__in=user_ids,
                is_approved=User.ApprovalStatus.APPROVED,
                role__in=[User.Role.STUDENT, User.Role.TEACHER],
            ).exclude(batch=batch)
            if not is_main_admin:
                # A sub-admin may only claim unassigned members — moving a
                # user out of another batch is a main-admin decision.
                candidates = candidates.filter(batch__isnull=True)
            updated = candidates.update(batch=batch)
            messages.success(request, f"تم إسناد {updated} عضواً إلى الدفعة")

        elif action == "assign_circles" and is_main_admin:
            circle_ids = request.POST.getlist("circle_ids")
            updated = Circle.objects.filter(
                pk__in=circle_ids, batch__isnull=True,
            ).update(batch=batch)
            messages.success(request, f"تم إسناد {updated} حلقة إلى الدفعة")

        elif action == "unassign_circle" and is_main_admin:
            circle = Circle.objects.filter(
                pk=request.POST.get("circle_id"), batch=batch,
            ).first()
            if circle:
                circle.batch = None
                circle.save(update_fields=["batch", "updated_at"])
                messages.success(request, f"تمت إزالة حلقة {circle.name} من الدفعة")

        elif action == "unassign_user":
            member = User.objects.filter(
                pk=request.POST.get("user_id"), batch=batch,
            ).first()
            if member:
                member.batch = None
                member.save(update_fields=["batch", "updated_at"])
                messages.success(request, f"تمت إزالة {member.full_name_ar} من الدفعة")

        elif action == "set_supervisors" and is_main_admin:
            supervisor_ids = request.POST.getlist("supervisor_ids")
            supervisors = User.objects.filter(
                pk__in=supervisor_ids,
                role=User.Role.SUB_ADMIN,
                is_approved=User.ApprovalStatus.APPROVED,
            )
            batch.sub_admin = supervisors.first()
            batch.save(update_fields=["sub_admin", "updated_at"])
            batch.sub_admins.set(supervisors)
            messages.success(request, "تم تحديث مشرفي الدفعة")

        elif action == "change_status" and is_main_admin:
            new_status = request.POST.get("status")
            if new_status in dict(Batch.Status.choices):
                batch.status = new_status
                batch.save(update_fields=["status", "updated_at"])
                messages.success(
                    request,
                    f"تم تغيير حالة الدفعة إلى {batch.get_status_display()}",
                )
            else:
                messages.error(request, "حالة غير صالحة")

        elif action == "delete" and is_main_admin:
            if batch.users.exists() or batch.circles.exists():
                messages.error(
                    request,
                    "لا يمكن حذف دفعة بها أعضاء أو حلقات — قم بأرشفتها بدلاً من ذلك",
                )
            else:
                name = batch.name
                batch.delete()
                messages.success(request, f"تم حذف الدفعة {name}")
                return redirect("accounts:admin_batch_list")

        return redirect("accounts:admin_batch_detail", pk=batch.pk)

    return render(
        request,
        "dashboard/admin/batches/detail.html",
        _batch_detail_context(request, batch),
    )


@login_required
@role_required(User.Role.MAIN_ADMIN)
def admin_batch_toggle_status(request, pk):
    """Explicit status change (POST only). Kept URL name for the list page's
    quick control; the blind GET cycle is gone."""
    if request.method != "POST":
        return redirect("accounts:admin_batch_list")
    batch = get_object_or_404(Batch, pk=pk)
    new_status = request.POST.get("status")
    if new_status not in dict(Batch.Status.choices):
        messages.error(request, "حالة غير صالحة")
        return redirect("accounts:admin_batch_list")
    batch.status = new_status
    batch.save(update_fields=["status", "updated_at"])
    messages.success(request, f"تم تغيير حالة الدفعة {batch.name} إلى {batch.get_status_display()}")
    return redirect("accounts:admin_batch_list")


@login_required
@role_required(User.Role.MAIN_ADMIN, User.Role.SUB_ADMIN)
def admin_batch_circles(request, pk):
    """Circles belonging to a specific batch — navigation step between
    Batch List → [Open Batch] → Halaqas → [Open] → supervisor_group_board."""
    batch = get_object_or_404(
        Batch.objects.select_related("sub_admin").prefetch_related("sub_admins"),
        pk=pk,
    )
    if (
        request.user.role != User.Role.MAIN_ADMIN
        and batch.sub_admin != request.user
        and not batch.sub_admins.filter(pk=request.user.pk).exists()
    ):
        raise PermissionDenied

    circles = (
        Circle.objects.filter(batch=batch)
        .select_related("teacher")
        .annotate(
            active_students=Count(
                "enrollments",
                filter=Q(enrollments__status=CircleEnrollment.Status.ACTIVE),
                distinct=True,
            ),
            sessions_count=Count("sessions", distinct=True),
        )
        .order_by("name")
    )

    return render(request, "dashboard/admin/batches/circles.html", {
        "batch": batch,
        "circles": circles,
    })
