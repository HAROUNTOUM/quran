import csv
import io
from datetime import date, datetime

from django import forms
from django.contrib import admin, messages
from django.contrib.admin.models import LogEntry
from django.contrib.auth.admin import UserAdmin as BaseUserAdmin
from django.core.exceptions import PermissionDenied
from django.db import transaction
from django.http import HttpResponse, HttpResponseRedirect
from django.template.response import TemplateResponse
from django.urls import path, reverse
from openpyxl import Workbook, load_workbook

from .models import StudentCard, TeacherAbsence, User


class AdminImportForm(forms.Form):
    file = forms.FileField(label="ملف الاستيراد")
    import_format = forms.ChoiceField(
        label="نوع الملف",
        choices=(("csv", "CSV"), ("xlsx", "Excel (XLSX)")),
        initial="csv",
    )


class ExportImportAdminMixin:
    change_list_template = "admin/import_export_change_list.html"
    export_fields = ()
    import_fields = ()
    import_lookup_field = None
    import_title = "استيراد البيانات"
    import_help_text = ""

    def get_urls(self):
        urls = super().get_urls()
        extra = [
            path(
                "import/",
                self.admin_site.admin_view(self.import_view),
                name=f"{self.model._meta.app_label}_{self.model._meta.model_name}_import",
            ),
        ]
        return extra + urls

    def import_url(self):
        return reverse(
            f"admin:{self.model._meta.app_label}_{self.model._meta.model_name}_import"
        )

    def export_value(self, obj, field_name):
        value = getattr(obj, field_name)
        if callable(value):
            value = value()
        if hasattr(value, "all"):
            return ", ".join(str(item) for item in value.all())
        if isinstance(value, (date, datetime)):
            return value.isoformat()
        return value

    def export_queryset_csv(self, queryset):
        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = (
            f'attachment; filename="{self.model._meta.model_name}.csv"'
        )
        writer = csv.writer(response)
        writer.writerow(self.export_fields)
        for obj in queryset:
            writer.writerow([self.export_value(obj, field) for field in self.export_fields])
        return response

    def export_queryset_xlsx(self, queryset):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = self.model._meta.verbose_name_plural[:31]
        worksheet.append(list(self.export_fields))
        for obj in queryset:
            worksheet.append([self.export_value(obj, field) for field in self.export_fields])

        buffer = io.BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        response = HttpResponse(
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = (
            f'attachment; filename="{self.model._meta.model_name}.xlsx"'
        )
        return response

    @admin.action(description="تصدير المحدد CSV")
    def export_selected_csv(self, request, queryset):
        return self.export_queryset_csv(queryset)

    @admin.action(description="تصدير المحدد Excel")
    def export_selected_xlsx(self, request, queryset):
        return self.export_queryset_xlsx(queryset)

    def parse_bool(self, value):
        if value is None or value == "":
            return None
        text = str(value).strip().lower()
        if text in {"1", "true", "yes", "y", "on", "نعم"}:
            return True
        if text in {"0", "false", "no", "n", "off", "لا"}:
            return False
        raise ValueError(f"قيمة منطقية غير صالحة: {value}")

    def parse_date(self, value):
        if not value:
            return None
        if isinstance(value, date):
            return value
        if isinstance(value, datetime):
            return value.date()
        text = str(value).strip()
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
            try:
                return datetime.strptime(text, fmt).date()
            except ValueError:
                continue
        raise ValueError(f"تاريخ غير صالح: {value}")

    def parse_uploaded_rows(self, uploaded_file, import_format):
        if import_format == "csv":
            decoded = io.StringIO(uploaded_file.read().decode("utf-8-sig"))
            reader = csv.DictReader(decoded)
            return list(reader)

        workbook = load_workbook(uploaded_file, data_only=True)
        worksheet = workbook.active
        rows = list(worksheet.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(header).strip() if header is not None else "" for header in rows[0]]
        return [dict(zip(headers, row)) for row in rows[1:]]

    def assign_import_value(self, obj, field_name, raw_value):
        if field_name.endswith("_date"):
            value = self.parse_date(raw_value)
        elif field_name.startswith("is_"):
            value = self.parse_bool(raw_value)
        else:
            value = raw_value
        setattr(obj, field_name, value)

    def import_row(self, row):
        raise NotImplementedError

    def import_view(self, request):
        if not self.has_add_permission(request) and not self.has_change_permission(request):
            raise PermissionDenied

        form = AdminImportForm(request.POST or None, request.FILES or None)
        context = {
            **self.admin_site.each_context(request),
            "title": self.import_title,
            "form": form,
            "opts": self.model._meta,
            "help_text": self.import_help_text,
            "fields": self.get_import_fields(),
        }

        if request.method == "POST" and form.is_valid():
            uploaded = form.cleaned_data["file"]
            import_format = form.cleaned_data["import_format"]
            rows = self.parse_uploaded_rows(uploaded, import_format)
            created = updated = errors = 0
            error_messages = []

            with transaction.atomic():
                for index, row in enumerate(rows, start=2):
                    try:
                        action = self.import_row(row)
                    except Exception as exc:
                        errors += 1
                        error_messages.append(f"السطر {index}: {exc}")
                        continue
                    if action == "created":
                        created += 1
                    else:
                        updated += 1

            if errors:
                self.message_user(
                    request,
                    f"تم استيراد {created} جديد و{updated} محدث، مع {errors} أخطاء.",
                    level=messages.WARNING,
                )
                for message in error_messages[:10]:
                    self.message_user(request, message, level=messages.ERROR)
            else:
                self.message_user(
                    request,
                    f"تم استيراد {created} جديد و{updated} محدث بنجاح.",
                    level=messages.SUCCESS,
                )
            return HttpResponseRedirect(
                reverse(
                    f"admin:{self.model._meta.app_label}_{self.model._meta.model_name}_changelist"
                )
            )

        return TemplateResponse(request, "admin/import_data.html", context)

    def get_import_fields(self):
        return self.import_fields or self.export_fields


@admin.register(User)
class UserAdmin(ExportImportAdminMixin, BaseUserAdmin):
    list_display = ["full_name_ar", "email", "role", "is_approved", "is_active", "created_at"]
    list_filter = ["role", "is_approved", "is_active", "is_staff", "gender"]
    search_fields = ["full_name_ar", "email", "phone", "username"]
    actions = [
        "approve_selected",
        "reject_selected",
        "activate_selected",
        "deactivate_selected",
        "export_selected_csv",
        "export_selected_xlsx",
    ]
    readonly_fields = ["created_at", "updated_at"]
    fieldsets = (
        (None, {"fields": ("username", "password")}),
        ("معلومات شخصية", {"fields": ("full_name_ar", "email", "phone", "gender")}),
        (
            "الصلاحيات",
            {
                "fields": (
                    "role",
                    "is_approved",
                    "rejection_reason",
                    "is_active",
                    "is_staff",
                    "is_superuser",
                    "groups",
                    "user_permissions",
                )
            },
        ),
        ("تواريخ مهمة", {"fields": ("last_login", "date_joined", "created_at", "updated_at")}),
    )
    add_fieldsets = (
        (
            None,
            {
                "classes": ("wide",),
                "fields": (
                    "username",
                    "email",
                    "full_name_ar",
                    "phone",
                    "gender",
                    "role",
                    "password1",
                    "password2",
                ),
            },
        ),
    )
    export_fields = (
        "username",
        "full_name_ar",
        "email",
        "phone",
        "gender",
        "role",
        "is_approved",
        "is_active",
        "is_staff",
        "is_superuser",
        "created_at",
    )
    import_fields = export_fields + ("rejection_reason",)
    import_lookup_field = "username"

    @admin.action(description="اعتماد المحدد")
    def approve_selected(self, request, queryset):
        updated = queryset.update(is_approved=User.ApprovalStatus.APPROVED, rejection_reason="")
        self.message_user(request, f"تم اعتماد {updated} مستخدمًا.", level=messages.SUCCESS)

    @admin.action(description="رفض المحدد")
    def reject_selected(self, request, queryset):
        updated = queryset.update(is_approved=User.ApprovalStatus.REJECTED)
        self.message_user(request, f"تم رفض {updated} مستخدمًا.", level=messages.SUCCESS)

    @admin.action(description="تفعيل المحدد")
    def activate_selected(self, request, queryset):
        updated = queryset.update(is_active=True)
        self.message_user(request, f"تم تفعيل {updated} مستخدمًا.", level=messages.SUCCESS)

    @admin.action(description="إلغاء تفعيل المحدد")
    def deactivate_selected(self, request, queryset):
        updated = queryset.update(is_active=False)
        self.message_user(request, f"تم إلغاء تفعيل {updated} مستخدمًا.", level=messages.SUCCESS)

    def import_row(self, row):
        lookup_value = row.get(self.import_lookup_field)
        if not lookup_value:
            raise ValueError("username مطلوب")

        obj, created = User.objects.get_or_create(username=lookup_value)
        if created:
            obj.set_unusable_password()
        for field_name in self.import_fields:
            if field_name == self.import_lookup_field or field_name not in row:
                continue
            self.assign_import_value(obj, field_name, row.get(field_name))
        obj.save()
        return "created" if created else "updated"


@admin.register(StudentCard)
class StudentCardAdmin(ExportImportAdminMixin, admin.ModelAdmin):
    list_display = ["card_number", "student", "created_at"]
    list_filter = ["created_at", "updated_at", "student__role"]
    search_fields = ["card_number", "student__full_name_ar", "student__username"]
    actions = ["export_selected_csv", "export_selected_xlsx"]
    export_fields = ("card_number", "student", "qr_code_data", "created_at")
    import_fields = ("card_number", "student_username", "qr_code_data")
    import_lookup_field = "card_number"

    def export_value(self, obj, field_name):
        if field_name == "student":
            return obj.student.username
        return super().export_value(obj, field_name)

    def import_row(self, row):
        lookup_value = row.get(self.import_lookup_field)
        if not lookup_value:
            raise ValueError("card_number مطلوب")

        obj, created = StudentCard.objects.get_or_create(card_number=lookup_value)
        student_username = row.get("student_username")
        if not student_username:
            raise ValueError("student_username مطلوب")
        obj.student = User.objects.get(username=student_username)
        if "qr_code_data" in row and row.get("qr_code_data") is not None:
            obj.qr_code_data = row.get("qr_code_data")
        obj.save()
        return "created" if created else "updated"


@admin.register(TeacherAbsence)
class TeacherAbsenceAdmin(ExportImportAdminMixin, admin.ModelAdmin):
    list_display = ["teacher", "start_date", "end_date", "duration_days", "status", "processed_by"]
    list_filter = ["status", "start_date", "end_date", "processed_by"]
    search_fields = [
        "teacher__full_name_ar",
        "teacher__username",
        "reason",
        "substitute_teacher__full_name_ar",
    ]
    actions = [
        "approve_selected",
        "reject_selected",
        "export_selected_csv",
        "export_selected_xlsx",
    ]
    export_fields = (
        "id",
        "teacher",
        "start_date",
        "end_date",
        "reason",
        "status",
        "rejection_reason",
        "substitute_teacher",
        "processed_by",
        "created_at",
    )
    import_fields = (
        "id",
        "teacher_username",
        "start_date",
        "end_date",
        "reason",
        "status",
        "rejection_reason",
        "substitute_teacher_username",
        "processed_by_username",
    )
    import_lookup_field = "id"

    def export_value(self, obj, field_name):
        if field_name == "teacher":
            return obj.teacher.username
        if field_name == "substitute_teacher":
            return obj.substitute_teacher.username if obj.substitute_teacher else ""
        if field_name == "processed_by":
            return obj.processed_by.username if obj.processed_by else ""
        return super().export_value(obj, field_name)

    @admin.action(description="اعتماد المحدد")
    def approve_selected(self, request, queryset):
        updated = queryset.update(status=TeacherAbsence.Status.APPROVED, rejection_reason="")
        self.message_user(request, f"تم اعتماد {updated} طلب غياب.", level=messages.SUCCESS)

    @admin.action(description="رفض المحدد")
    def reject_selected(self, request, queryset):
        updated = queryset.update(status=TeacherAbsence.Status.REJECTED)
        self.message_user(request, f"تم رفض {updated} طلب غياب.", level=messages.SUCCESS)

    def import_row(self, row):
        raw_id = row.get("id")
        obj = None
        created = True
        if raw_id:
            obj = TeacherAbsence.objects.filter(pk=raw_id).first()
            created = obj is None
        if obj is None:
            obj = TeacherAbsence()

        teacher_username = row.get("teacher_username")
        if not teacher_username:
            raise ValueError("teacher_username مطلوب")
        obj.teacher = User.objects.get(username=teacher_username)
        obj.start_date = self.parse_date(row.get("start_date"))
        obj.end_date = self.parse_date(row.get("end_date"))
        obj.reason = row.get("reason", "")
        obj.status = row.get("status") or TeacherAbsence.Status.PENDING
        obj.rejection_reason = row.get("rejection_reason", "")
        substitute_username = row.get("substitute_teacher_username")
        obj.substitute_teacher = (
            User.objects.get(username=substitute_username) if substitute_username else None
        )
        processed_by_username = row.get("processed_by_username")
        obj.processed_by = (
            User.objects.get(username=processed_by_username) if processed_by_username else None
        )
        obj.save()
        return "created" if created else "updated"


@admin.register(LogEntry)
class AuditLogAdmin(admin.ModelAdmin):
    date_hierarchy = "action_time"
    list_display = ["action_time", "user", "content_type", "object_repr", "action_flag"]
    list_filter = ["action_flag", "content_type", "user", "action_time"]
    search_fields = ["object_repr", "change_message", "user__username", "user__email"]
    readonly_fields = [
        "action_time",
        "user",
        "content_type",
        "object_id",
        "object_repr",
        "action_flag",
        "change_message",
    ]
    ordering = ["-action_time"]

    def has_add_permission(self, request):
        return False

    def has_change_permission(self, request, obj=None):
        return False

    def has_delete_permission(self, request, obj=None):
        return False
